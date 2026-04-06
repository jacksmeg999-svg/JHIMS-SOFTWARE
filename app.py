from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory
from functools import wraps
from datetime import datetime, date
import sqlite3, os, math, json, shutil, base64, io
import qrcode
from werkzeug.security import generate_password_hash, check_password_hash

APP_NAME = "DMH / JHIMS Hospital Web"
HOSPITAL_NAME = "JACKSON MUNICIPAL HOSPITAL"
HOSPITAL_ADDRESS = "P.O.BOX 49 DUNKWA-ON-OFFIN"
HOSPITAL_PHONE = ""
POWERED_BY = "Powered by JACKSTIDIOS @ {year}".format(year=datetime.now().year)

DB_PATH = os.path.join(os.path.dirname(__file__), "hospital_web.db")

DEFAULT_RECEIPT_PREFIX = "DMH"
DEFAULT_EMBALMING_FEE = 350.0
DEFAULT_DAILY_RATE = 16.0

BUILTIN_SERVICES = [
    "OPD","INVESTIGATION","PHARMACY","IPD","DRESSING","MORTUARY","OBS","DENTAL",
    "EYE","ULTRASOUND","ANC","OXYGEN","ENT","EMERGENCY","X-RAY","ECG"
]

BUILTIN_LAB_TESTS = [
    ("Full Blood Count (FBC)", 30.0, 50.0),
    ("ANC LABS", 80.0, 330.0),
    ("BF Malaria", 0.0, 20.0),
    ("UPT", 10.0, 20.0),
    ("LFT", 60.0, 80.0),
    ("RFT", 70.0, 90.0),
    ("BUE CR", 70.0, 90.0),
    ("LIPID PROFILE", 50.0, 70.0),
    ("HB", 0.0, 20.0),
    ("SICKLING", 0.0, 20.0),
    ("BLOOD GROUPING", 0.0, 20.0),
    ("G6PD", 0.0, 50.0),
    ("HBSAG", 0.0, 20.0),
    ("HCV", 0.0, 30.0),
    ("VDRL", 0.0, 20.0),
    ("TYPHOID", 30.0, 50.0),
    ("H.PYLORI", 30.0, 50.0),
    ("S.BILIRUBIN", 30.0, 50.0),
    ("URIC ACID", 30.0, 50.0),
    ("URINE R/E", 0.0, 20.0),
]

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(32))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

@app.context_processor
def inject_hospital():
    return dict(hospital=get_hospital())

def init_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS hospital_settings (
            id INTEGER PRIMARY KEY,
            hospital_name TEXT,
            address TEXT,
            phone TEXT,
            logo TEXT
        )
    """)
    conn.commit()
    conn.close()

def init_db():
    """Called once at startup to ensure all tables exist."""
    init_settings()
    init_ipd()

def init_ipd():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
CREATE TABLE IF NOT EXISTS ipd (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    patient_id TEXT,
    patient_name TEXT,
    age TEXT,
    sex TEXT,
    diagnosis TEXT,
    doctor TEXT,
    ward TEXT,
    bed TEXT,
    funding TEXT,
    admission_date TEXT,
    status TEXT DEFAULT 'Admitted'
)
""")

    c.execute("""
        CREATE TABLE IF NOT EXISTS ipd_services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admission_id INTEGER,
            service_name TEXT,
            amount REAL,
            date TEXT
        )
    """)

    c.execute("CREATE TABLE IF NOT EXISTS meta(key TEXT PRIMARY KEY, value TEXT)")

    c.execute(
        "CREATE TABLE IF NOT EXISTS users("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "username TEXT UNIQUE,"
        "password TEXT,"
        "role TEXT)"
    )

# ================= MORTUARY SETTINGS =================
    c.execute("""
        CREATE TABLE IF NOT EXISTS mortuary_settings(
        id INTEGER PRIMARY KEY,
        embalming_fee REAL,
        week1_rate REAL,
        week2_rate REAL,
        week3_rate REAL
    )
    """)

    c.execute("SELECT COUNT(*) FROM mortuary_settings")
    if c.fetchone()[0] == 0:
       c.execute("""
        INSERT INTO mortuary_settings 
        (embalming_fee, week1_rate, week2_rate, week3_rate)
        VALUES (350, 16, 13, 10)
    """)

    c.execute(
        "CREATE TABLE IF NOT EXISTS patient_records("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "date TEXT,"
        "receipt_number TEXT,"
        "patient_id TEXT,"
        "patient_name TEXT,"
        "service_received TEXT,"
        "amount_paid REAL,"
        "amount_in_words TEXT,"
        "cashier_name TEXT,"
        "payment_method TEXT,"
        "details TEXT)"
    )

    c.execute(
        "CREATE TABLE IF NOT EXISTS lab_requests("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "created_at TEXT,"
        "patient_id TEXT,"
        "patient_name TEXT,"
        "age TEXT,"
        "insured_status TEXT,"
        "tests_json TEXT,"
        "total_amount REAL,"
        "invoice_no TEXT,"
        "is_paid INTEGER DEFAULT 0,"
        "results_json TEXT,"
        "results_at TEXT)"
    )

    c.execute(
        "CREATE TABLE IF NOT EXISTS mortuary_cases("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "corpse_id TEXT UNIQUE,"
        "patient_id TEXT,"
        "corpse_name TEXT,"
        "sex TEXT,"
        "age TEXT,"
        "relative_name TEXT,"
        "relative_phone TEXT,"
        "deposit_date TEXT,"
        "release_date TEXT,"
        "days_total INTEGER,"
        "daily_rate REAL,"
        "embalming_fee REAL,"
        "extra_items TEXT,"
        "extra_amount REAL,"
        "total_amount REAL,"
        "status TEXT,"
        "is_paid INTEGER DEFAULT 0,"
        "brought_in_dead INTEGER DEFAULT 0,"
        "embalming_paid INTEGER DEFAULT 0,"
        "embalming_invoice TEXT,"
        "notes TEXT,"
        "cause_of_death TEXT,"
        "doctor_name TEXT)"
    )

    # IPD upgrade — beds and ward tables
    c.execute("""
        CREATE TABLE IF NOT EXISTS ipd_wards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ward_name TEXT UNIQUE,
            total_beds INTEGER DEFAULT 10
        )
    """)
    c.execute("SELECT COUNT(*) FROM ipd_wards")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO ipd_wards(ward_name,total_beds) VALUES(?,?)", [
            ("Male Ward", 20), ("Female Ward", 20), ("Children Ward", 10),
            ("Maternity Ward", 15), ("Emergency Ward", 10), ("Private Ward", 6),
        ])

    c.execute("""
        CREATE TABLE IF NOT EXISTS ipd_daily_charges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admission_id INTEGER,
            charge_date TEXT,
            description TEXT,
            amount REAL
        )
    """)

    c.execute(
        "CREATE TABLE IF NOT EXISTS lab_test_prices("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "name TEXT UNIQUE,"
        "insured_price REAL,"
        "non_insured_price REAL)"
    )

    c.execute(
        "CREATE TABLE IF NOT EXISTS services("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "name TEXT UNIQUE)"
    )

    c.execute(
        "CREATE TABLE IF NOT EXISTS activity_log("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "timestamp TEXT,"
        "username TEXT,"
        "action TEXT,"
        "details TEXT)"
    )

    # ================= RECORDS / OPD REGISTRATION =================
    c.execute("""
        CREATE TABLE IF NOT EXISTS opd_patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opd_number TEXT UNIQUE,
            patient_id TEXT UNIQUE,
            full_name TEXT,
            date_of_birth TEXT,
            age TEXT,
            sex TEXT,
            phone TEXT,
            address TEXT,
            next_of_kin TEXT,
            next_of_kin_phone TEXT,
            funding TEXT DEFAULT 'Non-Insured',
            nhis_number TEXT,
            ccc_code TEXT,
            consultation_fee REAL DEFAULT 0.0,
            consultation_paid INTEGER DEFAULT 0,
            registered_by TEXT,
            registered_at TEXT,
            status TEXT DEFAULT 'Active'
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS opd_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            opd_number TEXT,
            patient_id TEXT,
            visit_date TEXT,
            visit_type TEXT DEFAULT 'OPD',
            funding TEXT,
            nhis_number TEXT,
            ccc_code TEXT,
            consultation_fee REAL DEFAULT 0.0,
            consultation_paid INTEGER DEFAULT 0,
            triage_status TEXT DEFAULT 'Pending',
            doctor_status TEXT DEFAULT 'Pending',
            lab_status TEXT DEFAULT 'Pending',
            cashier_status TEXT DEFAULT 'Pending',
            pharmacy_status TEXT DEFAULT 'Pending',
            notes TEXT,
            serviced_by TEXT,
            serviced_at TEXT,
            current_location TEXT DEFAULT 'Records'
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS opd_settings (
            id INTEGER PRIMARY KEY,
            consultation_fee_non_insured REAL DEFAULT 10.0,
            opd_prefix TEXT DEFAULT 'OPD',
            patient_id_prefix TEXT DEFAULT 'JMH'
        )
    """)
    c.execute("SELECT COUNT(*) FROM opd_settings")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO opd_settings (consultation_fee_non_insured, opd_prefix, patient_id_prefix) VALUES (10.0, 'OPD', 'JMH')")

    # ── TRIAGE ──
    c.execute("""CREATE TABLE IF NOT EXISTS triage_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        visit_id INTEGER, patient_id TEXT, patient_name TEXT, opd_number TEXT,
        triage_date TEXT, triage_time TEXT, temperature REAL,
        bp_systolic INTEGER, bp_diastolic INTEGER, pulse INTEGER,
        respiration INTEGER, oxygen_sat REAL, weight REAL, height REAL, bmi REAL,
        chief_complaint TEXT, triage_level TEXT DEFAULT 'Normal',
        notes TEXT, triaged_by TEXT)""")

    # ── DOCTOR ──
    c.execute("""CREATE TABLE IF NOT EXISTS doctor_consultations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        visit_id INTEGER, patient_id TEXT, patient_name TEXT, opd_number TEXT,
        consult_date TEXT, doctor_name TEXT, chief_complaint TEXT,
        history TEXT, examination TEXT, diagnosis TEXT, treatment_plan TEXT,
        notes TEXT, status TEXT DEFAULT 'Open', follow_up_date TEXT)""")

    c.execute("""CREATE TABLE IF NOT EXISTS doctor_lab_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        consultation_id INTEGER, patient_id TEXT, patient_name TEXT,
        tests_json TEXT, requested_by TEXT, requested_at TEXT,
        status TEXT DEFAULT 'Pending')""")

    # ── PRESCRIPTIONS ──
    c.execute("""CREATE TABLE IF NOT EXISTS prescriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        consultation_id INTEGER, patient_id TEXT, patient_name TEXT,
        opd_number TEXT, prescribed_by TEXT, prescribed_at TEXT,
        drugs_json TEXT, total_amount REAL DEFAULT 0.0,
        invoice_no TEXT, status TEXT DEFAULT 'Pending',
        is_paid INTEGER DEFAULT 0, is_dispensed INTEGER DEFAULT 0, notes TEXT)""")

    # ── PHARMACY ──
    c.execute("""CREATE TABLE IF NOT EXISTS pharmacy_drugs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        drug_name TEXT UNIQUE, generic_name TEXT, category TEXT,
        unit TEXT DEFAULT 'Tablet', unit_price REAL DEFAULT 0.0,
        stock_qty INTEGER DEFAULT 0, reorder_level INTEGER DEFAULT 10,
        supplier TEXT, expiry_date TEXT, is_active INTEGER DEFAULT 1)""")

    c.execute("""CREATE TABLE IF NOT EXISTS pharmacy_dispensing (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        prescription_id INTEGER, patient_id TEXT, patient_name TEXT,
        dispensed_by TEXT, dispensed_at TEXT, drugs_json TEXT,
        total_amount REAL DEFAULT 0.0, invoice_no TEXT,
        is_paid INTEGER DEFAULT 0, receipt_no TEXT)""")

    # ── UNIVERSAL INVOICES ──
    c.execute("""CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_no TEXT UNIQUE, invoice_type TEXT,
        patient_id TEXT, patient_name TEXT, opd_number TEXT,
        items_json TEXT, total_amount REAL DEFAULT 0.0,
        created_at TEXT, created_by TEXT,
        is_paid INTEGER DEFAULT 0, paid_at TEXT,
        receipt_no TEXT, notes TEXT)""")

    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        default_users = [
            ("admin",    generate_password_hash("admin123"),    "Admin"),
            ("cashier",  generate_password_hash("cashier123"),  "Cashier"),
            ("lab",      generate_password_hash("lab123"),      "Lab"),
            ("mortuary", generate_password_hash("mortuary123"), "Mortuary"),
            ("reports",  generate_password_hash("reports123"),  "Reports"),
            ("records",  generate_password_hash("records123"),  "Records"),
            ("triage",   generate_password_hash("triage123"),   "Triage"),
            ("doctor",   generate_password_hash("doctor123"),   "Doctor"),
            ("pharmacy", generate_password_hash("pharmacy123"), "Pharmacy"),
        ]
        c.executemany("INSERT INTO users(username,password,role) VALUES(?,?,?)", default_users,)

    c.execute("SELECT COUNT(*) FROM lab_test_prices")
    if c.fetchone()[0] == 0:
        c.executemany(
            "INSERT INTO lab_test_prices(name,insured_price,non_insured_price) VALUES(?,?,?)",
            BUILTIN_LAB_TESTS,
        )

    c.execute("SELECT COUNT(*) FROM services")
    if c.fetchone()[0] == 0:
        c.executemany("INSERT INTO services(name) VALUES(?)", [(s,) for s in BUILTIN_SERVICES])

    conn.commit()
    conn.close()

def log_activity(username, action, details=""):
    conn = get_db()
    c = conn.cursor()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute(
        "INSERT INTO activity_log(timestamp,username,action,details) VALUES(?,?,?,?)",
        (ts, username, action, details),
    )
    conn.commit()
    conn.close()


def get_meta(key, default=None):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT value FROM meta WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row["value"] if row else default


def set_meta(key, value):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO meta(key,value) VALUES(?,?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, str(value)),
    )
    conn.commit()
    conn.close()


ONES = [
    "",
    "One",
    "Two",
    "Three",
    "Four",
    "Five",
    "Six",
    "Seven",
    "Eight",
    "Nine",
    "Ten",
    "Eleven",
    "Twelve",
    "Thirteen",
    "Fourteen",
    "Fifteen",
    "Sixteen",
    "Seventeen",
    "Eighteen",
    "Nineteen",
]
TENS = [
    "",
    "",
    "Twenty",
    "Thirty",
    "Forty",
    "Fifty",
    "Sixty",
    "Seventy",
    "Eighty",
    "Ninety",
]


def _chunk_to_words(n):
    parts = []
    if n >= 100:
        parts.append(ONES[n // 100])
        parts.append("Hundred")
        n %= 100
        if n:
            parts.append("and")
    if n >= 20:
        parts.append(TENS[n // 10])
        if n % 10:
            parts.append(ONES[n % 10])
    elif n > 0:
        parts.append(ONES[n])
    return " ".join(parts) if parts else "Zero"


def number_to_words(n):
    if n == 0:
        return "Zero"
    parts = []
    for label, unit in [("Billion", 1_000_000_000), ("Million", 1_000_000), ("Thousand", 1000)]:
        chunk = n // unit
        if chunk:
            parts.append(_chunk_to_words(chunk))
            parts.append(label)
            n %= unit
    if n:
        parts.append(_chunk_to_words(n))
    return " ".join(parts)


def amount_to_words(amount):
    if amount < 0:
        return "Minus " + amount_to_words(-amount)
    cedis = int(math.floor(amount))
    pesewas = int(round((amount - cedis) * 100))
    parts = []
    if cedis:
        parts.append(number_to_words(cedis))
        parts.append("Ghana Cedis")
    else:
        parts.append("Zero Ghana Cedis")
    if pesewas:
        parts.append("and")
        parts.append(number_to_words(pesewas))
        parts.append("Pesewas")
    parts.append("Only")
    return " ".join(parts)


def get_receipt_prefix():
    return get_meta("receipt_prefix", DEFAULT_RECEIPT_PREFIX) or DEFAULT_RECEIPT_PREFIX


def generate_receipt_number():
    prefix = get_receipt_prefix()
    last = get_meta("receipt_last", prefix + "/0000000") or prefix + "/0000000"
    if "/" in last:
        _, num = last.split("/", 1)
        sep = "/"
    else:
        num = last[len(prefix) :]
        sep = ""
    try:
        seq = int(num) + 1
        width = len(num)
    except Exception:
        seq = 1
        width = 7
    new = f"{seq:0{width}d}"
    receipt = f"{prefix}{sep}{new}"
    set_meta("receipt_last", receipt)
    return receipt


def generate_lab_invoice():
    last = get_meta("lab_inv_last", "INV000000") or "INV000000"
    prefix = "INV"
    num = last.replace(prefix, "")
    try:
        seq = int(num) + 1
        width = len(num)
    except Exception:
        seq = 1
        width = 6
    new = f"{seq:0{width}d}"
    inv = f"{prefix}{new}"
    set_meta("lab_inv_last", inv)
    return inv


def generate_invoice_no(prefix="INV"):
    key = f"inv_{prefix}_last"
    last = get_meta(key, f"{prefix}000000") or f"{prefix}000000"
    num = last.replace(prefix, "")
    try:
        seq = int(num) + 1
        width = max(len(num), 6)
    except Exception:
        seq = 1; width = 6
    inv = f"{prefix}{seq:0{width}d}"
    set_meta(key, inv)
    return inv

def generate_rx_no():   return generate_invoice_no("RX")
def generate_pharm_inv(): return generate_invoice_no("PH")



def generate_corpse_id():
    last = get_meta("corpse_last", "C0000") or "C0000"
    prefix = "".join(ch for ch in last if not ch.isdigit()) or "C"
    num = "".join(ch for ch in last if ch.isdigit()) or "0"
    try:
        seq = int(num) + 1
        width = max(len(num), 4)
    except Exception:
        seq = 1
        width = 4
    new = f"{seq:0{width}d}"
    cid = prefix + new
    set_meta("corpse_last", cid)
    return cid


def calculate_bill(deposit_date):
    settings = get_mortuary_settings()

    from datetime import datetime
    d1 = datetime.strptime(deposit_date, "%Y-%m-%d")
    d2 = datetime.today()

    days = (d2 - d1).days or 1

    total = 0

    for d in range(1, days+1):
        if d <= 7:
            total += settings["week1_rate"]
        elif d <= 14:
            total += settings["week2_rate"]
        else:
            total += settings["week3_rate"]

    return days, total

def get_hospital():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM hospital_settings LIMIT 1")
    data = c.fetchone()
    conn.close()
    return data

def login_required(role=None):
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "user" not in session:
                return redirect(url_for("login"))
            if role and session.get("role") not in (role, "Admin"):
                flash("Access denied for your role.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)

        return wrapper

    return deco


_db_initialized = False

@app.before_request
def startup():
    global _db_initialized
    if not _db_initialized:
        init_db()
        _db_initialized = True


@app.context_processor
def inject_globals():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, insured_price, non_insured_price FROM lab_test_prices ORDER BY name")
    tests = c.fetchall()
    c.execute("SELECT name FROM services ORDER BY name")
    services = [r["name"] for r in c.fetchall()]
    conn.close()
    return dict(
        APP_NAME=APP_NAME,
        HOSPITAL_NAME=HOSPITAL_NAME,
        HOSPITAL_ADDRESS=HOSPITAL_ADDRESS,
        HOSPITAL_PHONE=HOSPITAL_PHONE,
        POWERED_BY=POWERED_BY,
        LAB_TESTS=tests,
        SERVICES=services,
        now=datetime.now,
    )




def generate_qr(data: str) -> str:
    """Generate a QR code PNG as base64 string. Requires qrcode and pillow."""
    try:
        img = qrcode.make(data)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception:
        return ""

@app.route("/")
def index():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        user = c.fetchone()
        conn.close()
        if user and check_password_hash(user["password"], password):
            session["user"] = user["username"]
            session["role"] = user["role"]
            log_activity(user["username"], "LOGIN", "User logged in")
            return redirect(url_for("dashboard"))
        flash("Invalid username or password", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    if "user" in session:
        log_activity(session["user"], "LOGOUT", "User logged out")
    session.clear()
    return redirect(url_for("login"))




@app.route("/dashboard")
@login_required()
def dashboard():
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    c = conn.cursor()

    # Today's cash
    c.execute("SELECT COALESCE(SUM(amount_paid),0) FROM patient_records WHERE date=?", (today,))
    total_today = c.fetchone()[0]

    # Today's OPD visits
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=?", (today,))
    today_opd_count = c.fetchone()[0]

    # Today's NHIS
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=? AND funding='NHIS'", (today,))
    today_nhis = c.fetchone()[0]

    # Today's Non-Insured
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=? AND funding='Non-Insured'", (today,))
    today_noninsured = c.fetchone()[0]

    # Today's lab requests
    c.execute("SELECT COUNT(*) FROM lab_requests WHERE substr(created_at,1,10)=?", (today,))
    today_lab_count = c.fetchone()[0]

    # Active mortuary cases
    c.execute("SELECT COUNT(*) FROM mortuary_cases WHERE status='In-Mortuary'", )
    mortuary_count = c.fetchone()[0]

    # Active IPD patients
    try:
        c.execute("SELECT COUNT(*) FROM ipd WHERE status='Admitted'")
        ipd_count = c.fetchone()[0]
    except Exception:
        ipd_count = 0

    # Total receipts all time
    c.execute("SELECT COUNT(*) FROM patient_records")
    total_records = c.fetchone()[0]

    # Recent activity (Admin only)
    recent_activity = []
    c.execute("SELECT * FROM activity_log ORDER BY id DESC LIMIT 15")
    recent_activity = c.fetchall()

    # Patient history timeline
    pid = request.args.get("patient_id", "").strip()
    timeline = []
    if pid:
        c.execute("SELECT date, time('now') as t, 'Receipt' as source, receipt_number as ref, service_received as label, amount_paid as amount FROM patient_records WHERE patient_id=? ORDER BY id DESC LIMIT 20", (pid,))
        timeline += [dict(row) for row in c.fetchall()]
        c.execute("SELECT substr(created_at,1,10) as date, substr(created_at,12,8) as t, 'Lab' as source, invoice_no as ref, 'Lab invoice' as label, total_amount as amount FROM lab_requests WHERE patient_id=? ORDER BY id DESC LIMIT 20", (pid,))
        timeline += [dict(row) for row in c.fetchall()]
        c.execute("SELECT deposit_date as date, '00:00:00' as t, 'Mortuary' as source, corpse_id as ref, status as label, total_amount as amount FROM mortuary_cases WHERE patient_id=? ORDER BY id DESC LIMIT 20", (pid,))
        timeline += [dict(row) for row in c.fetchall()]
        timeline.sort(key=lambda item: (item.get("date") or "", item.get("t") or ""), reverse=True)

    conn.close()
    return render_template("dashboard.html",
        total_today=total_today, today=today,
        today_opd_count=today_opd_count,
        today_nhis=today_nhis,
        today_noninsured=today_noninsured,
        today_lab_count=today_lab_count,
        mortuary_count=mortuary_count,
        ipd_count=ipd_count,
        total_records=total_records,
        recent_activity=recent_activity,
        patient_id=pid, timeline=timeline)



@app.route("/cashier", methods=["GET", "POST"])
@login_required(role="Cashier")
def cashier():
    conn = get_db()
    c = conn.cursor()

    prefill = None

    # ================= GET PARAMETERS =================
    load_pid = request.args.get("load_pid", "").strip()
    corpse_id = request.args.get("corpse_id", "").strip()

    # ================= LOAD FROM MORTUARY (BY CORPSE ID) =================
    if corpse_id:
        c.execute("SELECT * FROM mortuary_cases WHERE corpse_id=?", (corpse_id,))
        m = c.fetchone()
        if m:
            prefill = {
                "patient_id": m["patient_id"],
                "patient_name": m["corpse_name"],
                "service_received": "MORTUARY",
                "amount_paid": m["embalming_fee"] if not m["embalming_paid"] else m["total_amount"],
                "details": f"{'Embalming fee' if not m['embalming_paid'] else 'Full mortuary bill'} for {m['corpse_name']} (Corpse ID: {m['corpse_id']})"
            }

    # ================= LOAD FROM LAB OR MORTUARY (BY PATIENT ID) =================
    elif load_pid:
        # First check OPD patients for name auto-fill
        c.execute("SELECT * FROM opd_patients WHERE patient_id=? OR opd_number=?", (load_pid, load_pid))
        opd_row = c.fetchone()

        # Check unpaid lab
        c.execute(
            "SELECT * FROM lab_requests WHERE patient_id=? AND is_paid=0 ORDER BY id DESC LIMIT 1",
            (load_pid,)
        )
        lab_row = c.fetchone()

        if lab_row:
            prefill = {
                "patient_id": lab_row["patient_id"],
                "patient_name": lab_row["patient_name"],
                "service_received": "INVESTIGATION",
                "amount_paid": lab_row["total_amount"],
                "details": f"LAB INVOICE {lab_row['invoice_no']}"
            }
        else:
            # Check unpaid mortuary
            c.execute(
                "SELECT * FROM mortuary_cases WHERE patient_id=? AND is_paid=0 ORDER BY id DESC LIMIT 1",
                (load_pid,)
            )
            m_row = c.fetchone()
            if m_row:
                prefill = {
                    "patient_id": m_row["patient_id"],
                    "patient_name": m_row["corpse_name"],
                    "service_received": "MORTUARY",
                    "amount_paid": m_row["embalming_fee"] if not m_row["embalming_paid"] else m_row["total_amount"],
                    "details": f"Mortuary case {m_row['corpse_id']}"
                }
            elif opd_row:
                # OPD patient found — prefill name/ID for new receipt
                prefill = {
                    "patient_id": opd_row["patient_id"],
                    "patient_name": opd_row["full_name"],
                    "service_received": "",
                    "amount_paid": "",
                    "details": f"OPD: {opd_row['opd_number']} | Funding: {opd_row['funding']}"
                }
                flash(f"Patient loaded: {opd_row['full_name']} ({opd_row['patient_id']})", "success")

        if not prefill and not opd_row:
            flash("No unpaid record found for that ID.", "danger")

    # ================= SAVE PAYMENT =================
    if request.method == "POST":
        patient_id = request.form.get("patient_id", "").strip()
        patient_name = request.form.get("patient_name", "").strip()
        service_received = request.form.get("service_received", "").strip()
        payment_method = request.form.get("payment_method", "").strip()
        amount_str = request.form.get("amount_paid", "0").strip()
        details = request.form.get("details", "").strip()

        if not patient_name or not service_received or not amount_str:
            flash("Fill patient, service and amount.", "danger")
        else:
            try:
                amount = float(amount_str)
            except:
                amount = 0.0

            receipt = generate_receipt_number()
            amount_words = amount_to_words(amount)
            today = datetime.now().strftime("%Y-%m-%d")
            cashier_name = session.get("user")

            c.execute(
                "INSERT INTO patient_records("
                "date,receipt_number,patient_id,patient_name,service_received,"
                "amount_paid,amount_in_words,cashier_name,payment_method,details"
                ") VALUES(?,?,?,?,?,?,?,?,?,?)",
                (
                    today,
                    receipt,
                    patient_id,
                    patient_name,
                    service_received,
                    amount,
                    amount_words,
                    cashier_name,
                    payment_method,
                    details,
                ),
            )

            # ================= MARK LAB AS PAID =================
            if service_received == "INVESTIGATION":
                c.execute(
                    "UPDATE lab_requests SET is_paid=1 WHERE patient_id=? AND is_paid=0",
                    (patient_id,)
                )

            # ================= MARK MORTUARY AS PAID =================
            if service_received == "Mortuary":
                c.execute(
                    "UPDATE mortuary_cases SET is_paid=1 WHERE patient_id=? AND total_amount > 0",
                    (patient_id,)
                )

            conn.commit()
            rec_id = c.lastrowid

            log_activity(
                cashier_name,
                "RECEIPT_CREATE",
                f"Receipt {receipt} for {patient_name}",
            )

            flash(f"Receipt saved: {receipt}", "success")
            conn.close()

            return redirect(url_for("receipt_view", rec_id=rec_id))

    # ================= LOAD RECEIPTS =================
    q = request.args.get("q", "").strip()

    if q:
        c.execute(
            "SELECT * FROM patient_records "
            "WHERE patient_name LIKE ? OR receipt_number LIKE ? OR patient_id LIKE ? "
            "ORDER BY id DESC LIMIT 100",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        )
    else:
        c.execute("SELECT * FROM patient_records ORDER BY id DESC LIMIT 50")

    rows = c.fetchall()
    conn.close()

    return render_template("cashier.html", rows=rows, q=q, prefill=prefill)

@app.route("/receipt/<int:rec_id>")

@login_required()
def receipt_view(rec_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM patient_records WHERE id=?", (rec_id,))
    rec = c.fetchone()
    conn.close()
    if not rec:
        flash("Receipt not found.", "danger")
        return redirect(url_for("cashier"))
    return render_template("receipt.html", rec=rec)


@app.route("/settings/hospital", methods=["GET", "POST"])
@login_required()
def hospital_settings():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        name = request.form["hospital_name"]
        address = request.form["address"]
        phone = request.form["phone"]
        logo = request.form["logo"]

        c.execute("DELETE FROM hospital_settings")
        c.execute("""
            INSERT INTO hospital_settings (hospital_name, address, phone, logo)
            VALUES (?, ?, ?, ?)
        """, (name, address, phone, logo))
        conn.commit()

    c.execute("SELECT * FROM hospital_settings LIMIT 1")
    data = c.fetchone()
    conn.close()

    return render_template("hospital_settings.html", data=data)


@app.route("/cashier/edit/<int:rec_id>", methods=["GET", "POST"])
@login_required(role="Admin")
def edit_receipt(rec_id):
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        patient_id = request.form.get("patient_id", "").strip()
        patient_name = request.form.get("patient_name", "").strip()
        service_received = request.form.get("service_received", "").strip()
        payment_method = request.form.get("payment_method", "").strip()
        amount_str = request.form.get("amount_paid", "0").strip()
        details = request.form.get("details", "").strip()
        try:
            amount = float(amount_str)
        except Exception:
            amount = 0.0
        amount_words = amount_to_words(amount)
        c.execute(
            "UPDATE patient_records SET patient_id=?,patient_name=?,service_received=?,"
            "amount_paid=?,amount_in_words=?,payment_method=?,details=? WHERE id=?",
            (
                patient_id,
                patient_name,
                service_received,
                amount,
                amount_words,
                payment_method,
                details,
                rec_id,
            ),
        )
        conn.commit()
        conn.close()
        flash("Receipt updated.", "success")
        return redirect(url_for("cashier"))
    c.execute("SELECT * FROM patient_records WHERE id=?", (rec_id,))
    rec = c.fetchone()
    conn.close()
    if not rec:
        flash("Receipt not found.", "danger")
        return redirect(url_for("cashier"))
    return render_template("edit_receipt.html", rec=rec)


@app.route("/cashier/delete/<int:rec_id>", methods=["POST"])
@login_required(role="Admin")
def delete_receipt(rec_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM patient_records WHERE id=?", (rec_id,))
    conn.commit()
    conn.close()
    flash("Receipt deleted.", "success")
    return redirect(url_for("cashier"))



@app.route("/lab", methods=["GET", "POST"])
@login_required(role="Lab")
def lab():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        patient_id = request.form.get("patient_id", "").strip()
        patient_name = request.form.get("patient_name", "").strip()
        age = request.form.get("age", "").strip()
        insured_status = request.form.get("insured_status", "Non-insured")
        tests_json = request.form.get("tests_json", "[]")
        total_str = request.form.get("total_amount", "0").strip()
        if not patient_name or tests_json == "[]":
            flash("Enter patient and select at least one test.", "danger")
        else:
            try:
                total = float(total_str)
            except Exception:
                total = 0.0
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            invoice = generate_lab_invoice()
            c.execute(
                "INSERT INTO lab_requests("
                "created_at,patient_id,patient_name,age,insured_status,"
                "tests_json,total_amount,invoice_no,is_paid,results_json,results_at"
                ") VALUES(?,?,?,?,?,?,?,?,0,NULL,NULL)",
                (
                    created_at,
                    patient_id,
                    patient_name,
                    age,
                    insured_status,
                    tests_json,
                    total,
                    invoice,
                ),
            )
            conn.commit()
            req_id = c.lastrowid
            log_activity(session.get("user"), "LAB_REQUEST", f"Invoice {invoice} for {patient_name}")
            conn.close()
            flash(f"Lab request saved. Invoice {invoice}", "success")
            return redirect(url_for("lab_invoice", req_id=req_id))

    # filters/search
    q = request.args.get("q", "").strip()
    today = datetime.now().strftime("%Y-%m-%d")
    if q:
        c.execute(
            "SELECT * FROM lab_requests WHERE patient_name LIKE ? OR patient_id LIKE ? OR invoice_no LIKE ? "
            "ORDER BY id DESC LIMIT 100",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        )
    else:
        c.execute("SELECT * FROM lab_requests ORDER BY id DESC LIMIT 50")
    rows = c.fetchall()

    # today's summary
    c.execute("SELECT COALESCE(SUM(total_amount),0) FROM lab_requests WHERE substr(created_at,1,10)=?", (today,))
    today_total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM lab_requests WHERE substr(created_at,1,10)=?", (today,))
    today_count = c.fetchone()[0]

    conn.close()
    return render_template("lab.html", rows=rows, q=q, today_total=today_total, today_count=today_count, today=today)


@app.route("/lab/invoice/<int:req_id>")
@login_required(role="Lab")
def lab_invoice(req_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM lab_requests WHERE id=?", (req_id,))
    req = c.fetchone()
    conn.close()
    if not req:
        flash("Lab request not found.", "danger")
        return redirect(url_for("lab"))
    try:
        tests = json.loads(req["tests_json"] or "[]")
    except Exception:
        tests = []
    return render_template("lab_invoice.html", req=req, tests=tests)


@app.route("/lab/results/<int:req_id>", methods=["GET", "POST"])
@login_required(role="Lab")
def lab_results(req_id):
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        results_json = request.form.get("results_json", "[]")
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute(
            "UPDATE lab_requests SET results_json=?,results_at=? WHERE id=?",
            (results_json, ts, req_id),
        )
        conn.commit()
        conn.close()
        flash("Lab results saved.", "success")
        return redirect(url_for("lab_results", req_id=req_id))
    c.execute("SELECT * FROM lab_requests WHERE id=?", (req_id,))
    req = c.fetchone()
    conn.close()
    if not req:
        flash("Lab request not found.", "danger")
        return redirect(url_for("lab"))
    try:
        tests = json.loads(req["tests_json"] or "[]")
    except Exception:
        tests = []
    try:
        results = json.loads(req["results_json"] or "[]")
    except Exception:
        results = []
    return render_template("lab_results.html", req=req, tests=tests, results=results)


@app.route("/lab/verify-payment", methods=["POST"])
@login_required(role="Lab")
def lab_verify_payment():
    pid = request.form.get("patient_id_verify", "").strip()
    if not pid:
        flash("Enter patient ID.", "danger")
        return redirect(url_for("lab"))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM patient_records WHERE patient_id=?", (pid,))
    count = c.fetchone()[0]
    conn.close()
    if count:
        flash(f"Payment found for Patient ID {pid}.", "success")
    else:
        flash(f"No payment yet for Patient ID {pid}.", "danger")
    return redirect(url_for("lab"))

def get_mortuary_settings():
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM mortuary_settings LIMIT 1")
    row = c.fetchone()

    conn.close()

    if row:
        return dict(row)
    else:
        # default values
        return {
            "embalming_fee": 350,
            "week1_rate": 16,
            "week2_rate": 13,
            "week3_rate": 10
        }


@app.route("/mortuary", methods=["GET", "POST"])
@login_required(role="Mortuary")
def mortuary():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        patient_id      = request.form.get("patient_id", "").strip()
        corpse_name     = request.form.get("corpse_name", "").strip()
        sex             = request.form.get("sex", "").strip()
        age             = request.form.get("age", "").strip()
        relative_name   = request.form.get("relative_name", "").strip()
        relative_phone  = request.form.get("relative_phone", "").strip()
        deposit_date    = request.form.get("deposit_date", "").strip()
        cause_of_death  = request.form.get("cause_of_death", "").strip()
        doctor_name     = request.form.get("doctor_name", "").strip()
        notes           = request.form.get("notes", "").strip()
        brought_in_dead = 1 if request.form.get("brought_in_dead") else 0

        if not corpse_name or not deposit_date:
            flash("Fill corpse name and deposit date.", "danger")
        else:
            # If patient ID provided, pull name from OPD patients
            if patient_id and not corpse_name:
                c.execute("SELECT full_name FROM opd_patients WHERE patient_id=?", (patient_id,))
                row = c.fetchone()
                if row:
                    corpse_name = row["full_name"]

            settings = get_mortuary_settings()
            embalming_fee = settings["embalming_fee"]
            cid = generate_corpse_id()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Generate embalming invoice number
            emb_inv = f"EMB/{cid}"

            c.execute("""
                INSERT INTO mortuary_cases(
                corpse_id, patient_id, corpse_name, sex, age,
                relative_name, relative_phone,
                deposit_date, release_date,
                days_total, daily_rate, embalming_fee,
                extra_items, extra_amount, total_amount,
                status, is_paid, brought_in_dead, embalming_paid,
                embalming_invoice, notes, cause_of_death, doctor_name
                ) VALUES (?,?,?,?,?,?,?,?,NULL,0,?,?,?,0,?,?,0,?,0,?,?,?,?)
            """, (
                cid, patient_id, corpse_name, sex, age,
                relative_name, relative_phone, deposit_date,
                settings["week1_rate"], embalming_fee,
                "", embalming_fee,
                "In-Mortuary", brought_in_dead, emb_inv,
                notes, cause_of_death, doctor_name
            ))
            conn.commit()
            log_activity(session.get("user"), "MORTUARY_REGISTER",
                         f"Registered {corpse_name} CID:{cid}")
            flash(f"✅ Body registered. Corpse ID: {cid} | Embalming Invoice: {emb_inv}", "success")
            flash(f"⚠️ Send patient ID {patient_id or cid} to Cashier to pay embalming fee of GHS {embalming_fee:.2f}", "warning")

    # FETCH with live bill
    c.execute("SELECT * FROM mortuary_cases ORDER BY id DESC")
    rows = c.fetchall()
    cases = []
    for r in rows:
        days, bill = calculate_mortuary_bill(r["deposit_date"])
        case = dict(r)
        case["live_days"] = days
        case["live_bill"] = bill
        cases.append(case)

    settings = get_mortuary_settings()
    conn.close()
    return render_template("mortuary.html", rows=cases, settings=settings)


@app.route("/mortuary/release-id", methods=["POST"])
@login_required(role="Mortuary")
def mortuary_release_by_id():
    cid = request.form.get("corpse_id_lookup", "").strip()

    if not cid:
        flash("Enter Corpse ID.", "danger")
        return redirect(url_for("mortuary"))

    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT id FROM mortuary_cases WHERE corpse_id=?", (cid,))
    row = c.fetchone()
    conn.close()

    if not row:
        flash("Corpse ID not found.", "danger")
        return redirect(url_for("mortuary"))

    return redirect(url_for("mortuary_invoice", case_id=row["id"]))


@app.route("/mortuary/release/<int:case_id>", methods=["GET", "POST"])
@login_required(role="Mortuary")
def mortuary_release(case_id):
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        release_date = request.form.get("release_date", "").strip()
        extra_items = request.form.get("extra_items", "").strip()
        extra_amount = request.form.get("extra_amount", "0").strip()
        c.execute("SELECT * FROM mortuary_cases WHERE id=?", (case_id,))
        m = c.fetchone()
        if not m:
            conn.close()
            flash("Case not found.", "danger")
            return redirect(url_for("mortuary"))
        days, total = calculate_mortuary_bill(m["deposit_date"])
        daily_rate = float(get_meta("mort_daily_rate", DEFAULT_DAILY_RATE) or DEFAULT_DAILY_RATE)
        embalming_fee = float(
            get_meta("embalming_fee", DEFAULT_EMBALMING_FEE) or DEFAULT_EMBALMING_FEE
        )
        try:
            extra = float(extra_amount or 0)
        except Exception:
            extra = 0.0
        c.execute(
            "UPDATE mortuary_cases SET release_date=?,days_total=?,daily_rate=?,embalming_fee=?,"
            "extra_items=?,extra_amount=?,total_amount=?,status=? WHERE id=?",
            (
                release_date,
                days,
                daily_rate,
                embalming_fee,
                extra_items,
                extra,
                total,
                "Released",
                case_id,
            ),
        )
        conn.commit()
        conn.close()
        flash("Release data saved and bill updated.", "success")
        return redirect(url_for("mortuary_invoice", case_id=case_id))
    c.execute("SELECT * FROM mortuary_cases WHERE id=?", (case_id,))
    case = c.fetchone()
    conn.close()
    if not case:
        flash("Case not found.", "danger")
        return redirect(url_for("mortuary"))
    return render_template("mortuary_release.html", case=case)


@app.route("/mortuary/invoice/<int:case_id>")
@login_required(role="Mortuary")
def mortuary_invoice(case_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM mortuary_cases WHERE id=?", (case_id,))
    case = c.fetchone()
    conn.close()
    if not case:
        flash("Case not found.", "danger")
        return redirect(url_for("mortuary"))
    return render_template("mortuary_invoice.html", case=case)


def reports_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("Admin", "Reports", "Cashier"):
            flash("Access denied to reports.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return wrapper

# =========================
# VIEW
# =========================
@app.route("/mortuary/view/<int:id>")
@login_required(role="Mortuary")
def mortuary_view(id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM mortuary_cases WHERE id=?", (id,))
    r = c.fetchone()
    conn.close()

    if not r:
        return "Record not found"

    return render_template("mortuary_view.html", r=r)

# =========================
# DELETE
# =========================
@app.route("/mortuary/delete/<int:id>")
@login_required(role="Mortuary")
def delete_mortuary(id):
    conn = get_db()
    conn.execute("DELETE FROM mortuary_cases WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect("/mortuary")


@app.route("/reports")
@reports_required
def reports_home():
    return render_template("reports_home.html")


@app.route("/reports/patient", methods=["GET", "POST"])
@reports_required
def reports_patient():
    rows = []
    total = 0.0
    start = end = None
    if request.method == "POST":
        start = request.form.get("start")
        end = request.form.get("end")
        if start and end:
            conn = get_db()
            c = conn.cursor()
            c.execute(
                "SELECT * FROM patient_records WHERE date BETWEEN ? AND ? ORDER BY date,id",
                (start, end),
            )
            rows = c.fetchall()
            c.execute(
                "SELECT COALESCE(SUM(amount_paid),0) FROM patient_records WHERE date BETWEEN ? AND ?",
                (start, end),
            )
            total = c.fetchone()[0]
            conn.close()
    return render_template("reports_patient.html", rows=rows, total=total, start=start, end=end)


@app.route("/reports/services", methods=["GET", "POST"])
@reports_required
def reports_services():
    rows = []
    grand = 0.0
    start = end = None
    if request.method == "POST":
        start = request.form.get("start")
        end = request.form.get("end")
        if start and end:
            conn = get_db()
            c = conn.cursor()
            c.execute(
                "SELECT service_received,COUNT(*) as count_rec, SUM(amount_paid) as total_amount "
                "FROM patient_records WHERE date BETWEEN ? AND ? "
                "GROUP BY service_received ORDER BY service_received",
                (start, end),
            )
            rows = c.fetchall()
            c.execute(
                "SELECT COALESCE(SUM(amount_paid),0) FROM patient_records WHERE date BETWEEN ? AND ?",
                (start, end),
            )
            grand = c.fetchone()[0]
            conn.close()
    return render_template(
        "reports_services.html", rows=rows, grand=grand, start=start, end=end
    )


@app.route("/reports/mortuary", methods=["GET", "POST"])
@reports_required
def reports_mortuary():
    rows = []
    total = 0.0
    start = end = None
    if request.method == "POST":
        start = request.form.get("start")
        end = request.form.get("end")
        if start and end:
            conn = get_db()
            c = conn.cursor()
            c.execute(
                "SELECT * FROM mortuary_cases WHERE deposit_date BETWEEN ? AND ? "
                "ORDER BY deposit_date,id",
                (start, end),
            )
            rows = c.fetchall()
            c.execute(
                "SELECT COALESCE(SUM(total_amount),0) FROM mortuary_cases WHERE deposit_date BETWEEN ? AND ?",
                (start, end),
            )
            total = c.fetchone()[0]
            conn.close()
    return render_template("reports_mortuary.html", rows=rows, total=total, start=start, end=end)

@app.route("/mortuary/settings", methods=["GET","POST"])
@login_required(role="Admin")
def mortuary_settings():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        embalming = request.form["embalming_fee"]
        w1 = request.form["week1"]
        w2 = request.form["week2"]
        w3 = request.form["week3"]

        c.execute("""
        UPDATE mortuary_settings
        SET embalming_fee=?, week1_rate=?, week2_rate=?, week3_rate=?
        """, (embalming, w1, w2, w3))

        conn.commit()

    c.execute("SELECT * FROM mortuary_settings LIMIT 1")
    s = c.fetchone()
    conn.close()

    return render_template("mortuary_settings.html", s=s)


def calculate_mortuary_bill(deposit_date):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM mortuary_settings LIMIT 1")
    s = c.fetchone()

    week1 = s["week1_rate"]
    week2 = s["week2_rate"]
    week3 = s["week3_rate"]

    d1 = datetime.strptime(deposit_date, "%Y-%m-%d")
    d2 = datetime.today()

    days = (d2 - d1).days
    if days <= 0:
        days = 1

    total = 0

    for day in range(1, days + 1):
        if day <= 7:
            total += week1
        elif day <= 14:
            total += week2
        else:
            total += week3

    conn.close()
    return days, total


@app.route("/admin/users", methods=["GET", "POST"])
@login_required(role="Admin")
def manage_users():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "").strip() or "Cashier"
        if username and password:
            try:
                c.execute(
                    "INSERT INTO users(username,password,role) VALUES(?,?,?)",
                    (username, generate_password_hash(password), role),
                )
                conn.commit()
                flash("User added.", "success")
                log_activity(session.get("user"), "USER_ADD", f"Added {username}")
            except sqlite3.IntegrityError:
                flash("Username already exists.", "danger")
        else:
            flash("Enter username and password.", "danger")
    c.execute("SELECT * FROM users ORDER BY username")
    users = c.fetchall()
    conn.close()
    return render_template("manage_users.html", users=users)


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required(role="Admin")
def edit_user(user_id):
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "").strip()
        if password:
            c.execute("UPDATE users SET password=?,role=? WHERE id=?", (generate_password_hash(password), role, user_id))
        else:
            c.execute("UPDATE users SET role=? WHERE id=?", (role, user_id))
        conn.commit()
        conn.close()
        flash("User updated.", "success")
        return redirect(url_for("manage_users"))
    c.execute("SELECT * FROM users WHERE id=?", (user_id,))
    user = c.fetchone()
    conn.close()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("manage_users"))
    return render_template("edit_user.html", user=user)


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required(role="Admin")
def delete_user(user_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    flash("User deleted.", "success")
    return redirect(url_for("manage_users"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required()
def change_password():
    if request.method == "POST":
        current = request.form.get("current", "").strip()
        new = request.form.get("new", "").strip()
        if not new:
            flash("Enter new password.", "danger")
            return redirect(url_for("change_password"))
        conn = get_db()
        c = conn.cursor()
        c.execute(
            "SELECT * FROM users WHERE username=?",
            (session.get("user"),),
        )
        row = c.fetchone()
        if not row or not check_password_hash(row["password"], current):
            conn.close()
            flash("Current password incorrect.", "danger")
            return redirect(url_for("change_password"))
        c.execute(
            "UPDATE users SET password=? WHERE username=?", (generate_password_hash(new), session.get("user"))
        )
        conn.commit()
        conn.close()
        flash("Password changed.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/settings/receipt", methods=["GET", "POST"])
@login_required(role="Admin")
def settings_receipt():
    if request.method == "POST":
        prefix = request.form.get("prefix", "").strip() or DEFAULT_RECEIPT_PREFIX
        last = request.form.get("last", "").strip()
        set_meta("receipt_prefix", prefix)
        if last:
            set_meta("receipt_last", last)
        flash("Receipt settings saved.", "success")
    prefix = get_receipt_prefix()
    last = get_meta("receipt_last", prefix + "/0000000")
    return render_template("settings_receipt.html", prefix=prefix, last=last)



@app.route("/settings/theme", methods=["GET", "POST"])
@login_required(role="Admin")
def settings_theme():
    if request.method == "POST":
        theme = request.form.get("theme", "light")
        if theme not in ("light", "luxury"):
            theme = "light"
        set_meta("ui_theme", theme)
        flash("Theme updated.", "success")
    current = get_meta("ui_theme", "light") or "light"
    return render_template("settings_theme.html", current=current)
@app.route("/settings/lab-prices", methods=["GET", "POST"])
@login_required(role="Admin")
def settings_lab_prices():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        for key, val in request.form.items():
            if key.startswith("ins_") or key.startswith("non_"):
                field, sid = key.split("_", 1)
                try:
                    price = float(val or 0)
                except Exception:
                    price = 0.0
                if field == "ins":
                    c.execute(
                        "UPDATE lab_test_prices SET insured_price=? WHERE id=?",
                        (price, sid),
                    )
                else:
                    c.execute(
                        "UPDATE lab_test_prices SET non_insured_price=? WHERE id=?",
                        (price, sid),
                    )
        conn.commit()
        flash("Lab prices updated.", "success")
    c.execute("SELECT * FROM lab_test_prices ORDER BY name")
    rows = c.fetchall()
    conn.close()
    return render_template("settings_lab_prices.html", rows=rows)


@app.route("/settings/services", methods=["GET", "POST"])
@login_required(role="Admin")
def settings_services():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name = request.form.get("name", "").strip()
            if name:
                try:
                    c.execute("INSERT INTO services(name) VALUES(?)", (name,))
                    conn.commit()
                    flash("Service added.", "success")
                except sqlite3.IntegrityError:
                    flash("Service already exists.", "danger")
        elif action == "delete":
            sid = request.form.get("service_id")
            c.execute("DELETE FROM services WHERE id=?", (sid,))
            conn.commit()
            flash("Service deleted.", "success")
    c.execute("SELECT * FROM services ORDER BY name")
    rows = c.fetchall()
    conn.close()
    return render_template("settings_services.html", rows=rows)

@app.route("/settings", methods=["GET", "POST"])
@login_required(role="Admin")
def settings():
    conn = get_db()
    c = conn.cursor()

    if request.method == "POST":
        name = request.form.get("hospital_name")
        address = request.form.get("address")
        phone = request.form.get("phone")
        logo = request.form.get("logo")

        c.execute("SELECT * FROM hospital_settings LIMIT 1")
        existing = c.fetchone()

        if existing:
            c.execute("""
                UPDATE hospital_settings
                SET hospital_name=?, address=?, phone=?, logo=?
                WHERE id=?
            """, (name, address, phone, logo, existing["id"]))
        else:
            c.execute("""
                INSERT INTO hospital_settings (hospital_name, address, phone, logo)
                VALUES (?, ?, ?, ?)
            """, (name, address, phone, logo))

        conn.commit()
        flash("Settings updated successfully!", "success")

    c.execute("SELECT * FROM hospital_settings LIMIT 1")
    data = c.fetchone()

    conn.close()
    return render_template("settings.html", data=data)


@app.route("/admin/activity")
@login_required(role="Admin")
def activity_log():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM activity_log ORDER BY id DESC LIMIT 200")
    rows = c.fetchall()
    conn.close()
    return render_template("activity_log.html", rows=rows)



def bill_consolidation_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        if session.get("role") not in ("Admin", "Cashier"):
            flash("Access denied to bill consolidation.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper

@app.route("/bill-consolidation", methods=["GET", "POST"])
@bill_consolidation_required
def bill_consolidation():
    pid = None
    receipts = []
    labs = []
    morts = []
    total_receipts = total_labs = total_morts = 0.0
    if request.method == "POST":
        pid = request.form.get("patient_id", "").strip()
    else:
        pid = request.args.get("patient_id", "").strip()
    if pid:
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT * FROM patient_records WHERE patient_id=? ORDER BY date,id", (pid,))
        receipts = c.fetchall()
        c.execute("SELECT * FROM lab_requests WHERE patient_id=? ORDER BY id", (pid,))
        labs = c.fetchall()
        c.execute("SELECT * FROM mortuary_cases WHERE patient_id=? ORDER BY id", (pid,))
        morts = c.fetchall()
        c.execute("SELECT COALESCE(SUM(amount_paid),0) FROM patient_records WHERE patient_id=?", (pid,))
        total_receipts = c.fetchone()[0]
        c.execute("SELECT COALESCE(SUM(total_amount),0) FROM lab_requests WHERE patient_id=?", (pid,))
        total_labs = c.fetchone()[0]
        c.execute("SELECT COALESCE(SUM(total_amount),0) FROM mortuary_cases WHERE patient_id=?", (pid,))
        total_morts = c.fetchone()[0]
        conn.close()
    grand = (total_receipts or 0) + (total_labs or 0) + (total_morts or 0)
    return render_template(
        "bill_consolidation.html",
        patient_id=pid,
        receipts=receipts,
        labs=labs,
        morts=morts,
        total_receipts=total_receipts,
        total_labs=total_labs,
        total_morts=total_morts,
        grand=grand,
    )

@app.route("/admin/backup")
@login_required(role="Admin")
def admin_backup():
    if not os.path.exists(DB_PATH):
        flash("Database not found.", "danger")
        return redirect(url_for("dashboard"))
    backup_dir = os.path.join(os.path.dirname(__file__), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"hospital_web_{ts}.db"
    dest = os.path.join(backup_dir, fname)
    shutil.copy2(DB_PATH, dest)
    flash("Backup created: " + fname, "success")
    return send_from_directory(backup_dir, fname, as_attachment=True)

@app.route("/mortuary/generate-bill", methods=["POST"])
@login_required(role="Mortuary")
def generate_mortuary_bill():
    print("GENERATE BILL CLICKED")
    conn = get_db()
    c = conn.cursor()

    corpse_id = request.form.get("corpse_id")

    c.execute("SELECT * FROM mortuary_cases WHERE corpse_id=?", (corpse_id,))
    case = c.fetchone()

    if not case:
        flash("Invalid Corpse ID", "danger")
        return redirect(url_for("mortuary"))

    settings = get_mortuary_settings()

    days, storage_bill = calculate_mortuary_bill(case["deposit_date"])

    embalming_fee = settings["embalming_fee"]

    extra = float(request.form.get("extra_amount") or 0)

    total = storage_bill + embalming_fee + extra

    conn.close()

    return render_template(
        "mortuary.html",
        case=case,
        days=days,
        storage_bill=storage_bill,
        embalming_fee=embalming_fee,
        extra=extra,
        total=total,
        show_bill=True
    )


# ============================================================

def generate_opd_number():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT opd_prefix FROM opd_settings LIMIT 1")
    row = c.fetchone()
    prefix = row["opd_prefix"] if row else "OPD"
    conn.close()
    last = get_meta("opd_last", f"{prefix}/000000") or f"{prefix}/000000"
    if "/" in last:
        _, num = last.split("/", 1)
    else:
        num = "000000"
    try:
        seq = int(num) + 1
        width = max(len(num), 6)
    except Exception:
        seq = 1
        width = 6
    new_num = f"{seq:0{width}d}"
    opd_no = f"{prefix}/{new_num}"
    set_meta("opd_last", opd_no)
    return opd_no


def generate_patient_id():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT patient_id_prefix FROM opd_settings LIMIT 1")
    row = c.fetchone()
    prefix = row["patient_id_prefix"] if row else "JMH"
    conn.close()
    last = get_meta("pid_last", f"{prefix}000000") or f"{prefix}000000"
    num_part = last.replace(prefix, "")
    try:
        seq = int(num_part) + 1
        width = max(len(num_part), 6)
    except Exception:
        seq = 1
        width = 6
    new_pid = f"{prefix}{seq:0{width}d}"
    set_meta("pid_last", new_pid)
    return new_pid


def generate_ccc_code(nhis_number):
    import hashlib
    raw = f"NHIS-{nhis_number}-{datetime.now().strftime('%Y%m%d')}"
    h = hashlib.md5(raw.encode()).hexdigest().upper()
    return f"CCC/{h[:4]}-{h[4:8]}-{h[8:12]}"


def get_opd_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_settings LIMIT 1")
    row = c.fetchone()
    conn.close()
    if row:
        return dict(row)
    return {"consultation_fee_non_insured": 10.0, "opd_prefix": "OPD", "patient_id_prefix": "JMH"}


@app.route("/records")
@login_required(role="Records")
def records():
    conn = get_db()
    c = conn.cursor()
    q = request.args.get("q", "").strip()
    today = datetime.now().strftime("%Y-%m-%d")
    if q:
        c.execute("""
            SELECT * FROM opd_patients
            WHERE full_name LIKE ? OR patient_id LIKE ? OR opd_number LIKE ? OR phone LIKE ? OR nhis_number LIKE ?
            ORDER BY id DESC LIMIT 100
        """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"))
    else:
        c.execute("SELECT * FROM opd_patients ORDER BY id DESC LIMIT 60")
    patients = c.fetchall()
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=?", (today,))
    today_visits = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM opd_patients WHERE substr(registered_at,1,10)=?", (today,))
    today_new = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=? AND funding='NHIS'", (today,))
    today_nhis = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM opd_visits WHERE visit_date=? AND funding='Non-Insured'", (today,))
    today_noninsured = c.fetchone()[0]
    conn.close()
    settings = get_opd_settings()
    return render_template("records.html",
        patients=patients, q=q, today=today,
        today_visits=today_visits, today_new=today_new,
        today_nhis=today_nhis, today_noninsured=today_noninsured,
        settings=settings)


@app.route("/records/register", methods=["GET", "POST"])
@login_required(role="Records")
def records_register():
    settings = get_opd_settings()
    if request.method == "POST":
        full_name   = request.form.get("full_name", "").strip()
        dob         = request.form.get("date_of_birth", "").strip()
        age         = request.form.get("age", "").strip()
        sex         = request.form.get("sex", "").strip()
        phone       = request.form.get("phone", "").strip()
        address     = request.form.get("address", "").strip()
        nok         = request.form.get("next_of_kin", "").strip()
        nok_phone   = request.form.get("next_of_kin_phone", "").strip()
        funding     = request.form.get("funding", "Non-Insured").strip()
        nhis_number = request.form.get("nhis_number", "").strip()
        if not full_name or not sex:
            flash("Full name and sex are required.", "danger")
            return render_template("records_register.html", settings=settings)
        ccc_code = ""
        if funding == "NHIS" and nhis_number:
            ccc_code = generate_ccc_code(nhis_number)
        consultation_fee = 0.0
        if funding == "Non-Insured":
            consultation_fee = float(settings.get("consultation_fee_non_insured", 10.0))
        opd_number = generate_opd_number()
        patient_id = generate_patient_id()
        now_str    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today      = datetime.now().strftime("%Y-%m-%d")
        cashier    = session.get("user")
        conn = get_db()
        c = conn.cursor()
        c.execute("""
            INSERT INTO opd_patients
            (opd_number, patient_id, full_name, date_of_birth, age, sex, phone, address,
             next_of_kin, next_of_kin_phone, funding, nhis_number, ccc_code,
             consultation_fee, consultation_paid, registered_by, registered_at, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,0,?,?,'Active')
        """, (opd_number, patient_id, full_name, dob, age, sex, phone, address,
              nok, nok_phone, funding, nhis_number, ccc_code,
              consultation_fee, cashier, now_str))
        c.execute("""
            INSERT INTO opd_visits
            (opd_number, patient_id, visit_date, visit_type, funding, nhis_number, ccc_code,
             consultation_fee, consultation_paid, serviced_by, serviced_at, current_location)
            VALUES (?,?,?,'New Registration',?,?,?,?,0,?,?,'Records')
        """, (opd_number, patient_id, today, funding, nhis_number,
              ccc_code, consultation_fee, cashier, now_str))
        visit_id = c.lastrowid
        conn.commit()
        conn.close()
        log_activity(cashier, "OPD_REGISTER", f"New patient {full_name} OPD:{opd_number} PID:{patient_id}")
        flash(f"Patient registered! OPD No: {opd_number}  |  Patient ID: {patient_id}", "success")
        if funding == "Non-Insured":
            flash(f"Consultation fee of GHS {consultation_fee:.2f} must be paid at Cashier.", "warning")
            return redirect(url_for("records_send_to_cashier", visit_id=visit_id))
        else:
            return redirect(url_for("records_visit_detail", visit_id=visit_id))
    return render_template("records_register.html", settings=settings)


@app.route("/records/service/<opd_number>", methods=["GET", "POST"])
@login_required(role="Records")
def records_service_card(opd_number):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_patients WHERE opd_number=? OR patient_id=?", (opd_number, opd_number))
    patient = c.fetchone()
    conn.close()
    if not patient:
        flash("Patient not found.", "danger")
        return redirect(url_for("records"))
    settings = get_opd_settings()
    if request.method == "POST":
        funding     = request.form.get("funding", patient["funding"]).strip()
        nhis_number = request.form.get("nhis_number", patient["nhis_number"] or "").strip()
        today       = datetime.now().strftime("%Y-%m-%d")
        now_str     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cashier     = session.get("user")
        ccc_code = ""
        if funding == "NHIS" and nhis_number:
            ccc_code = generate_ccc_code(nhis_number)
        consultation_fee = 0.0
        if funding == "Non-Insured":
            consultation_fee = float(settings.get("consultation_fee_non_insured", 10.0))
        conn = get_db()
        c = conn.cursor()
        c.execute("""
            INSERT INTO opd_visits
            (opd_number, patient_id, visit_date, visit_type, funding, nhis_number, ccc_code,
             consultation_fee, consultation_paid, serviced_by, serviced_at, current_location)
            VALUES (?,?,?,'Re-Visit',?,?,?,?,0,?,?,'Records')
        """, (patient["opd_number"], patient["patient_id"], today,
              funding, nhis_number, ccc_code, consultation_fee, cashier, now_str))
        visit_id = c.lastrowid
        c.execute("UPDATE opd_patients SET funding=?, nhis_number=?, ccc_code=? WHERE id=?",
                  (funding, nhis_number, ccc_code, patient["id"]))
        conn.commit()
        conn.close()
        log_activity(cashier, "OPD_SERVICE_CARD", f"Card serviced {patient['full_name']} OPD:{patient['opd_number']}")
        flash(f"Card serviced for {patient['full_name']}.", "success")
        if funding == "Non-Insured":
            flash(f"Consultation fee GHS {consultation_fee:.2f} required.", "warning")
            return redirect(url_for("records_send_to_cashier", visit_id=visit_id))
        else:
            return redirect(url_for("records_visit_detail", visit_id=visit_id))
    return render_template("records_service_card.html", patient=patient, settings=settings)


@app.route("/records/visit/<int:visit_id>")
@login_required(role="Records")
def records_visit_detail(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    if not visit:
        conn.close()
        flash("Visit not found.", "danger")
        return redirect(url_for("records"))
    c.execute("SELECT * FROM opd_patients WHERE opd_number=?", (visit["opd_number"],))
    patient = c.fetchone()
    conn.close()
    return render_template("records_visit_detail.html", visit=visit, patient=patient)


@app.route("/records/send-cashier/<int:visit_id>")
@login_required(role="Records")
def records_send_to_cashier(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE opd_visits SET current_location='Cashier', cashier_status='Pending' WHERE id=?", (visit_id,))
    conn.commit()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    conn.close()
    log_activity(session.get("user"), "OPD_SEND_CASHIER", f"Visit {visit_id} sent to Cashier")
    flash("Patient sent to Cashier for consultation fee.", "info")
    if visit:
        return redirect(url_for("cashier") + f"?load_pid={visit['patient_id']}")
    return redirect(url_for("records"))


@app.route("/records/send-triage/<int:visit_id>")
@login_required(role="Records")
def records_send_to_triage(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    if not visit:
        conn.close()
        flash("Visit not found.", "danger")
        return redirect(url_for("records"))
    if visit["funding"] == "Non-Insured" and visit["consultation_paid"] == 0:
        conn.close()
        flash("Cannot send to Triage. Consultation fee not yet paid.", "danger")
        return redirect(url_for("records_visit_detail", visit_id=visit_id))
    c.execute("UPDATE opd_visits SET current_location='Triage', triage_status='Pending' WHERE id=?", (visit_id,))
    conn.commit()
    conn.close()
    log_activity(session.get("user"), "OPD_SEND_TRIAGE", f"Visit {visit_id} sent to Triage")
    flash("Patient sent to Triage.", "success")
    return redirect(url_for("records_visit_detail", visit_id=visit_id))


@app.route("/records/send-doctor/<int:visit_id>")
@login_required(role="Records")
def records_send_to_doctor(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE opd_visits SET current_location='Doctor', doctor_status='Pending' WHERE id=?", (visit_id,))
    conn.commit()
    conn.close()
    log_activity(session.get("user"), "OPD_SEND_DOCTOR", f"Visit {visit_id} sent to Doctor")
    flash("Patient sent to Doctor.", "success")
    return redirect(url_for("records_visit_detail", visit_id=visit_id))


@app.route("/records/send-lab/<int:visit_id>")
@login_required(role="Records")
def records_send_to_lab(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    c.execute("UPDATE opd_visits SET current_location='Lab', lab_status='Pending' WHERE id=?", (visit_id,))
    conn.commit()
    conn.close()
    log_activity(session.get("user"), "OPD_SEND_LAB", f"Visit {visit_id} sent to Lab")
    flash("Patient sent to Lab.", "success")
    if visit:
        return redirect(url_for("lab") + f"?q={visit['patient_id']}")
    return redirect(url_for("records"))


@app.route("/records/visit/<int:visit_id>/mark-paid", methods=["POST"])
@login_required(role="Cashier")
def records_mark_consultation_paid(visit_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE opd_visits SET consultation_paid=1, cashier_status='Paid', current_location='Triage' WHERE id=?", (visit_id,))
    conn.commit()
    conn.close()
    log_activity(session.get("user"), "OPD_CONSULT_PAID", f"Consultation paid visit {visit_id}")
    flash("Consultation fee paid. Patient cleared for Triage.", "success")
    return redirect(url_for("records_visit_detail", visit_id=visit_id))


@app.route("/records/patient/<patient_id>")
@login_required(role="Records")
def records_patient_history(patient_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_patients WHERE patient_id=? OR opd_number=?", (patient_id, patient_id))
    patient = c.fetchone()
    if not patient:
        conn.close()
        flash("Patient not found.", "danger")
        return redirect(url_for("records"))
    c.execute("SELECT * FROM opd_visits WHERE patient_id=? ORDER BY id DESC", (patient["patient_id"],))
    visits = c.fetchall()
    c.execute("SELECT * FROM patient_records WHERE patient_id=? ORDER BY id DESC", (patient["patient_id"],))
    receipts = c.fetchall()
    c.execute("SELECT * FROM lab_requests WHERE patient_id=? ORDER BY id DESC", (patient["patient_id"],))
    labs = c.fetchall()
    conn.close()
    return render_template("records_patient_history.html",
                           patient=patient, visits=visits, receipts=receipts, labs=labs)


@app.route("/records/today")
@login_required(role="Records")
def records_today():
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT v.*, p.full_name, p.phone, p.sex, p.age
        FROM opd_visits v
        JOIN opd_patients p ON v.patient_id = p.patient_id
        WHERE v.visit_date=?
        ORDER BY v.id DESC
    """, (today,))
    visits = c.fetchall()
    conn.close()
    return render_template("records_today.html", visits=visits, today=today)


@app.route("/records/settings", methods=["GET", "POST"])
@login_required(role="Admin")
def records_settings():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        fee    = request.form.get("consultation_fee_non_insured", "10.0")
        prefix = request.form.get("opd_prefix", "OPD").strip()
        pidpfx = request.form.get("patient_id_prefix", "JMH").strip()
        try:
            fee = float(fee)
        except Exception:
            fee = 10.0
        c.execute("UPDATE opd_settings SET consultation_fee_non_insured=?, opd_prefix=?, patient_id_prefix=?",
                  (fee, prefix, pidpfx))
        conn.commit()
        flash("Records settings saved.", "success")
    c.execute("SELECT * FROM opd_settings LIMIT 1")
    s = c.fetchone()
    conn.close()
    return render_template("records_settings.html", s=s)


# ============================================================
#  MORTUARY UPGRADES
# ============================================================

@app.route("/mortuary/edit/<int:case_id>", methods=["GET", "POST"])
@login_required(role="Mortuary")
def mortuary_edit(case_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM mortuary_cases WHERE id=?", (case_id,))
    case = c.fetchone()
    if not case:
        conn.close()
        flash("Case not found.", "danger")
        return redirect(url_for("mortuary"))
    if request.method == "POST":
        corpse_name    = request.form.get("corpse_name","").strip()
        sex            = request.form.get("sex","").strip()
        age            = request.form.get("age","").strip()
        relative_name  = request.form.get("relative_name","").strip()
        relative_phone = request.form.get("relative_phone","").strip()
        deposit_date   = request.form.get("deposit_date","").strip()
        cause_of_death = request.form.get("cause_of_death","").strip()
        doctor_name    = request.form.get("doctor_name","").strip()
        notes          = request.form.get("notes","").strip()
        c.execute("""UPDATE mortuary_cases SET corpse_name=?,sex=?,age=?,relative_name=?,
            relative_phone=?,deposit_date=?,cause_of_death=?,doctor_name=?,notes=?
            WHERE id=?""",
            (corpse_name,sex,age,relative_name,relative_phone,deposit_date,
             cause_of_death,doctor_name,notes,case_id))
        conn.commit()
        conn.close()
        log_activity(session.get("user"),"MORTUARY_EDIT",f"Edited case {case_id}")
        flash("Record updated.", "success")
        return redirect(url_for("mortuary_view", id=case_id))
    conn.close()
    return render_template("mortuary_edit.html", case=case)


@app.route("/mortuary/mark-embalming-paid/<int:case_id>", methods=["POST"])
@login_required(role="Cashier")
def mortuary_mark_embalming_paid(case_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE mortuary_cases SET embalming_paid=1 WHERE id=?", (case_id,))
    conn.commit()
    conn.close()
    log_activity(session.get("user"),"MORTUARY_EMBALMING_PAID",f"Embalming paid case {case_id}")
    flash("Embalming fee marked as paid. Daily billing now active.", "success")
    return redirect(url_for("mortuary_view", id=case_id))


@app.route("/mortuary/discharge/<int:id>", methods=["GET", "POST"])
@login_required(role="Mortuary")
def discharge(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM mortuary_cases WHERE id=?", (id,))
    r = c.fetchone()
    if not r:
        conn.close()
        flash("Record not found.", "danger")
        return redirect(url_for("mortuary"))
    if r["embalming_paid"] == 0:
        conn.close()
        flash("⚠️ Cannot discharge. Embalming fee not yet paid.", "danger")
        return redirect(url_for("mortuary_view", id=id))
    if request.method == "POST":
        extra_items  = request.form.get("extra_items","").strip()
        extra_amount = request.form.get("extra_amount","0").strip()
        release_date = request.form.get("release_date", datetime.now().strftime("%Y-%m-%d")).strip()
        try:
            extra = float(extra_amount or 0)
        except Exception:
            extra = 0.0
        days, storage_bill = calculate_mortuary_bill(r["deposit_date"])
        settings = get_mortuary_settings()
        total = settings["embalming_fee"] + storage_bill + extra
        c.execute("""UPDATE mortuary_cases SET release_date=?,days_total=?,daily_rate=?,
            extra_items=?,extra_amount=?,total_amount=?,status=?,is_paid=0
            WHERE id=?""",
            (release_date, days, settings["week1_rate"],
             extra_items, extra, total, "Awaiting Payment", id))
        conn.commit()
        conn.close()
        log_activity(session.get("user"),"MORTUARY_DISCHARGE_BILL",f"Discharge bill generated case {id}")
        flash(f"Discharge bill generated. Total: GHS {total:.2f}. Send to Cashier for payment.", "success")
        return redirect(url_for("mortuary_invoice", case_id=id))
    conn.close()
    return render_template("mortuary_discharge.html", case=r)


@app.route("/mortuary/reports", methods=["GET","POST"])
@login_required(role="Mortuary")
def mortuary_reports():
    conn = get_db()
    c = conn.cursor()
    rows = []
    total_amount = 0.0
    start = end = None
    status_filter = request.form.get("status_filter","") if request.method=="POST" else ""
    if request.method == "POST":
        start = request.form.get("start","")
        end   = request.form.get("end","")
        if start and end:
            if status_filter:
                c.execute("SELECT * FROM mortuary_cases WHERE deposit_date BETWEEN ? AND ? AND status=? ORDER BY id DESC",
                          (start, end, status_filter))
            else:
                c.execute("SELECT * FROM mortuary_cases WHERE deposit_date BETWEEN ? AND ? ORDER BY id DESC",
                          (start, end))
            rows = c.fetchall()
            c.execute("SELECT COALESCE(SUM(total_amount),0) FROM mortuary_cases WHERE deposit_date BETWEEN ? AND ?",
                      (start, end))
            total_amount = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM mortuary_cases WHERE status='In-Mortuary'")
    in_mortuary = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM mortuary_cases WHERE status='Released'")
    released = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM mortuary_cases WHERE status='Awaiting Payment'")
    awaiting = c.fetchone()[0]
    conn.close()
    return render_template("mortuary_reports.html", rows=rows, total_amount=total_amount,
        start=start, end=end, status_filter=status_filter,
        in_mortuary=in_mortuary, released=released, awaiting=awaiting)


# ============================================================
#  IPD UPGRADES
# ============================================================

@app.route("/ipd")
@login_required()
def ipd():
    conn = get_db()
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT COUNT(*) FROM ipd WHERE status='Admitted'")
    admitted = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM ipd")
    total = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM ipd WHERE date(admission_date)=?", (today,))
    today_admissions = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM ipd WHERE status='Discharged'")
    discharged = c.fetchone()[0]

    # Ward occupancy
    c.execute("SELECT ward, COUNT(*) as cnt FROM ipd WHERE status='Admitted' GROUP BY ward")
    ward_counts = {r["ward"]: r["cnt"] for r in c.fetchall()}
    c.execute("SELECT * FROM ipd_wards")
    wards = c.fetchall()
    conn.close()
    return render_template("ipd.html", admitted=admitted, total=total,
        today_admissions=today_admissions, discharged=discharged,
        ward_counts=ward_counts, wards=wards)


@app.route('/ipd/admit', methods=['GET', 'POST'])
@login_required()
def ipd_admit():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM ipd_wards")
    wards = c.fetchall()
    if request.method == 'POST':
        patient_id     = request.form.get('patient_id','').strip()
        patient_name   = request.form.get('patient_name','').strip()
        age            = request.form.get('age','').strip()
        sex            = request.form.get('sex','').strip()
        diagnosis      = request.form.get('diagnosis','').strip()
        doctor         = request.form.get('doctor','').strip()
        ward           = request.form.get('ward','').strip()
        bed            = request.form.get('bed','').strip()
        funding        = request.form.get('funding','Non-Insured').strip()
        nhis_number    = request.form.get('nhis_number','').strip()
        admission_date = request.form.get('admission_date', datetime.now().strftime("%Y-%m-%d")).strip()
        # Auto-fill name from OPD if patient_id given
        if patient_id and not patient_name:
            c.execute("SELECT full_name FROM opd_patients WHERE patient_id=?", (patient_id,))
            row = c.fetchone()
            if row:
                patient_name = row["full_name"]
        if not patient_name:
            flash("Patient name is required.", "danger")
            conn.close()
            return render_template("ipd_admit.html", wards=wards)
        c.execute("""INSERT INTO ipd
            (patient_id,patient_name,age,sex,diagnosis,doctor,ward,bed,funding,admission_date,status)
            VALUES(?,?,?,?,?,?,?,?,?,?,'Admitted')""",
            (patient_id,patient_name,age,sex,diagnosis,doctor,ward,bed,funding,admission_date))
        adm_id = c.lastrowid
        # Auto-add daily bed charge
        c.execute("INSERT INTO ipd_daily_charges(admission_id,charge_date,description,amount) VALUES(?,?,?,?)",
                  (adm_id, admission_date, "Bed/Ward fee", 50.0))
        conn.commit()
        conn.close()
        log_activity(session.get("user"),"IPD_ADMIT",f"Admitted {patient_name}")
        flash(f"Patient admitted successfully.", "success")
        return redirect(url_for("ipd_patients"))
    conn.close()
    return render_template("ipd_admit.html", wards=wards)


@app.route('/ipd/patients')
@login_required()
def ipd_patients():
    conn = get_db()
    c = conn.cursor()
    q = request.args.get("q","").strip()
    status_f = request.args.get("status","Admitted").strip()
    if q:
        c.execute("SELECT * FROM ipd WHERE (patient_name LIKE ? OR patient_id LIKE ?) ORDER BY id DESC",
                  (f"%{q}%",f"%{q}%"))
    elif status_f == "All":
        c.execute("SELECT * FROM ipd ORDER BY id DESC")
    else:
        c.execute("SELECT * FROM ipd WHERE status=? ORDER BY id DESC", (status_f,))
    patients = c.fetchall()
    conn.close()
    return render_template("ipd_patients.html", patients=patients, q=q, status_f=status_f)


@app.route('/ipd/patient/<int:id>', methods=["GET","POST"])
@login_required()
def ipd_patient_profile(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM ipd WHERE id=?", (id,))
    patient = c.fetchone()
    if not patient:
        conn.close()
        flash("Patient not found.","danger")
        return redirect(url_for("ipd_patients"))
    # Daily charges
    c.execute("SELECT * FROM ipd_daily_charges WHERE admission_id=? ORDER BY id DESC", (id,))
    charges = c.fetchall()
    total_charges = sum(ch["amount"] for ch in charges)
    # Lab requests
    c.execute("SELECT * FROM lab_requests WHERE patient_id=? ORDER BY id DESC LIMIT 10", (patient["patient_id"],))
    labs = c.fetchall()
    # Receipts
    c.execute("SELECT * FROM patient_records WHERE patient_id=? ORDER BY id DESC LIMIT 10", (patient["patient_id"],))
    receipts = c.fetchall()
    conn.close()
    return render_template("ipd_profile.html", patient=patient, charges=charges,
                           total_charges=total_charges, labs=labs, receipts=receipts)


@app.route('/ipd/add-charge/<int:id>', methods=["POST"])
@login_required()
def ipd_add_charge(id):
    description = request.form.get("description","").strip()
    amount = request.form.get("amount","0").strip()
    try: amount = float(amount)
    except: amount = 0.0
    charge_date = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO ipd_daily_charges(admission_id,charge_date,description,amount) VALUES(?,?,?,?)",
              (id, charge_date, description, amount))
    conn.commit()
    conn.close()
    flash("Charge added.", "success")
    return redirect(url_for("ipd_patient_profile", id=id))


@app.route('/ipd/discharge/<int:id>', methods=["GET","POST"])
@login_required()
def discharge_patient(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM ipd WHERE id=?", (id,))
    patient = c.fetchone()
    if not patient:
        conn.close()
        flash("Patient not found.","danger")
        return redirect(url_for("ipd_patients"))
    c.execute("SELECT COALESCE(SUM(amount),0) FROM ipd_daily_charges WHERE admission_id=?", (id,))
    total_charges = c.fetchone()[0]
    if request.method == "POST":
        discharge_date = request.form.get("discharge_date", datetime.now().strftime("%Y-%m-%d"))
        discharge_notes = request.form.get("discharge_notes","").strip()
        c.execute("UPDATE ipd SET status='Discharged' WHERE id=?", (id,))
        conn.commit()
        conn.close()
        log_activity(session.get("user"),"IPD_DISCHARGE",f"Discharged {patient['patient_name']}")
        flash(f"Patient discharged. Total bill: GHS {total_charges:.2f}. Send to Cashier.", "success")
        return redirect(url_for("ipd_bill", id=id))
    conn.close()
    return render_template("ipd_discharge.html", patient=patient, total_charges=total_charges)


@app.route('/ipd/bill/<int:id>')
@login_required()
def ipd_bill(id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM ipd WHERE id=?", (id,))
    patient = c.fetchone()
    c.execute("SELECT * FROM ipd_daily_charges WHERE admission_id=? ORDER BY charge_date", (id,))
    charges = c.fetchall()
    total = sum(ch["amount"] for ch in charges)
    conn.close()
    return render_template("ipd_bill.html", patient=patient, charges=charges, total=total)


# ============================================================
#  UNIFIED PRICE SETTINGS (ADMIN)
# ============================================================

@app.route("/admin/price-settings", methods=["GET","POST"])
@login_required(role="Admin")
def admin_price_settings():
    conn = get_db()
    c = conn.cursor()
    if request.method == "POST":
        section = request.form.get("section","")
        if section == "mortuary":
            c.execute("UPDATE mortuary_settings SET embalming_fee=?,week1_rate=?,week2_rate=?,week3_rate=?",
                (request.form.get("embalming_fee",350),
                 request.form.get("week1_rate",16),
                 request.form.get("week2_rate",13),
                 request.form.get("week3_rate",10)))
            flash("Mortuary prices updated.", "success")
        elif section == "opd":
            fee = request.form.get("consultation_fee",10)
            try: fee = float(fee)
            except: fee = 10.0
            c.execute("UPDATE opd_settings SET consultation_fee_non_insured=?", (fee,))
            flash("OPD consultation fee updated.", "success")
        elif section == "lab":
            for key, val in request.form.items():
                if key.startswith("ins_") or key.startswith("non_"):
                    field, sid = key.split("_",1)
                    try: price = float(val or 0)
                    except: price = 0.0
                    if field == "ins":
                        c.execute("UPDATE lab_test_prices SET insured_price=? WHERE id=?", (price,sid))
                    else:
                        c.execute("UPDATE lab_test_prices SET non_insured_price=? WHERE id=?", (price,sid))
            flash("Lab prices updated.", "success")
        conn.commit()
    c.execute("SELECT * FROM mortuary_settings LIMIT 1")
    mort_s = c.fetchone()
    c.execute("SELECT * FROM opd_settings LIMIT 1")
    opd_s = c.fetchone()
    c.execute("SELECT * FROM lab_test_prices ORDER BY name")
    lab_prices = c.fetchall()
    conn.close()
    return render_template("admin_price_settings.html",
        mort_s=mort_s, opd_s=opd_s, lab_prices=lab_prices)


@app.route("/api/patient-name")
def api_patient_name():
    pid = request.args.get("pid","").strip()
    if not pid:
        return {"name": None}
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM opd_patients WHERE patient_id=? OR opd_number=?", (pid, pid))
    row = c.fetchone()
    conn.close()
    if row:
        return {"name": row["full_name"], "patient_id": row["patient_id"],
                "age": row["age"], "sex": row["sex"], "nhis": row["nhis_number"],
                "funding": row["funding"]}
    return {"name": None}



# ============================================================
#  UNIVERSAL LOOKUP API
# ============================================================

@app.route("/api/lookup")
def api_lookup():
    q = request.args.get("q","").strip()
    if not q:
        return {"found": False}
    conn = get_db(); c = conn.cursor()
    result = {"found": False}
    c.execute("SELECT * FROM opd_patients WHERE patient_id=? OR opd_number=?", (q,q))
    row = c.fetchone()
    if row:
        conn.close()
        return {"found":True,"patient_id":row["patient_id"],"name":row["full_name"],
                "opd_number":row["opd_number"],"funding":row["funding"],
                "age":row["age"] or "","sex":row["sex"] or "","nhis":row["nhis_number"] or ""}
    c.execute("SELECT * FROM mortuary_cases WHERE corpse_id=?", (q,))
    row = c.fetchone()
    if row:
        conn.close()
        return {"found":True,"patient_id":row["corpse_id"],"name":row["corpse_name"],"opd_number":"","funding":"Mortuary","type":"mortuary"}
    c.execute("SELECT * FROM invoices WHERE invoice_no=?", (q,))
    row = c.fetchone()
    if row:
        r = dict(row); conn.close()
        return {"found":True,"patient_id":r["patient_id"],"name":r["patient_name"],"opd_number":r.get("opd_number","") or "","funding":"","invoice_no":r["invoice_no"],"total":r["total_amount"]}
    conn.close()
    return {"found":False}


@app.route("/api/pending-invoices")
@login_required(role="Cashier")
def api_pending_invoices():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM invoices WHERE is_paid=0 ORDER BY id DESC LIMIT 50")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"invoices": rows, "count": len(rows)}


# ============================================================
#  TRIAGE MODULE
# ============================================================

@app.route("/triage")
@login_required(role="Triage")
def triage():
    conn = get_db(); c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("""SELECT v.*, p.full_name, p.age, p.sex, p.funding
        FROM opd_visits v JOIN opd_patients p ON v.patient_id=p.patient_id
        WHERE v.visit_date=? AND v.triage_status='Pending' ORDER BY v.id ASC""", (today,))
    pending = c.fetchall()
    c.execute("SELECT * FROM triage_records WHERE triage_date=? ORDER BY id DESC", (today,))
    done = c.fetchall()
    c.execute("SELECT COUNT(*) FROM triage_records WHERE triage_date=?", (today,))
    today_count = c.fetchone()[0]
    conn.close()
    return render_template("triage.html", pending=pending, done=done, today=today, today_count=today_count)


@app.route("/triage/record/<int:visit_id>", methods=["GET","POST"])
@login_required(role="Triage")
def triage_record(visit_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    if not visit:
        conn.close(); flash("Visit not found.","danger"); return redirect(url_for("triage"))
    c.execute("SELECT * FROM opd_patients WHERE patient_id=?", (visit["patient_id"],))
    patient = c.fetchone()
    if request.method == "POST":
        temp=request.form.get("temperature",""); bp_s=request.form.get("bp_systolic","")
        bp_d=request.form.get("bp_diastolic",""); pulse=request.form.get("pulse","")
        resp=request.form.get("respiration",""); o2=request.form.get("oxygen_sat","")
        weight=request.form.get("weight",""); height=request.form.get("height","")
        complaint=request.form.get("chief_complaint","").strip()
        level=request.form.get("triage_level","Normal"); notes=request.form.get("notes","").strip()
        today=datetime.now().strftime("%Y-%m-%d"); ttime=datetime.now().strftime("%H:%M")
        bmi=None
        try:
            w=float(weight); h=float(height)/100
            if h>0: bmi=round(w/(h*h),1)
        except: pass
        c.execute("""INSERT INTO triage_records
            (visit_id,patient_id,patient_name,opd_number,triage_date,triage_time,
             temperature,bp_systolic,bp_diastolic,pulse,respiration,oxygen_sat,
             weight,height,bmi,chief_complaint,triage_level,notes,triaged_by)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (visit_id,visit["patient_id"],patient["full_name"] if patient else "",
             visit["opd_number"],today,ttime,
             temp or None,bp_s or None,bp_d or None,pulse or None,
             resp or None,o2 or None,weight or None,height or None,bmi,
             complaint,level,notes,session.get("user")))
        c.execute("UPDATE opd_visits SET triage_status='Done', current_location='Doctor', doctor_status='Pending' WHERE id=?", (visit_id,))
        conn.commit(); conn.close()
        log_activity(session.get("user"),"TRIAGE",f"Triaged {patient['full_name'] if patient else visit_id}")
        flash("Vitals saved. Patient sent to Doctor.","success")
        return redirect(url_for("triage"))
    conn.close()
    return render_template("triage_record.html", visit=visit, patient=patient)


# ============================================================
#  DOCTOR MODULE
# ============================================================

@app.route("/doctor")
@login_required(role="Doctor")
def doctor():
    conn = get_db(); c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("""SELECT v.*, p.full_name, p.age, p.sex, p.funding, p.nhis_number
        FROM opd_visits v JOIN opd_patients p ON v.patient_id=p.patient_id
        WHERE v.visit_date=? AND v.doctor_status='Pending' ORDER BY v.id ASC""", (today,))
    waiting = c.fetchall()
    c.execute("SELECT * FROM doctor_consultations WHERE consult_date=? ORDER BY id DESC", (today,))
    consultations = c.fetchall()
    c.execute("SELECT COUNT(*) FROM doctor_consultations WHERE consult_date=?", (today,))
    today_count = c.fetchone()[0]
    conn.close()
    return render_template("doctor.html", waiting=waiting, consultations=consultations, today=today, today_count=today_count)


@app.route("/doctor/consult/<int:visit_id>", methods=["GET","POST"])
@login_required(role="Doctor")
def doctor_consult(visit_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM opd_visits WHERE id=?", (visit_id,))
    visit = c.fetchone()
    if not visit:
        conn.close(); flash("Visit not found.","danger"); return redirect(url_for("doctor"))
    c.execute("SELECT * FROM opd_patients WHERE patient_id=?", (visit["patient_id"],))
    patient = c.fetchone()
    c.execute("SELECT * FROM triage_records WHERE visit_id=? ORDER BY id DESC LIMIT 1", (visit_id,))
    vitals = c.fetchone()
    c.execute("SELECT * FROM doctor_consultations WHERE patient_id=? ORDER BY id DESC LIMIT 5", (visit["patient_id"],))
    prev_consults = c.fetchall()
    c.execute("SELECT * FROM lab_test_prices ORDER BY name")
    lab_tests = c.fetchall()
    c.execute("SELECT * FROM pharmacy_drugs WHERE is_active=1 ORDER BY drug_name")
    drugs_list = c.fetchall()
    if request.method == "POST":
        action=request.form.get("action","save")
        complaint=request.form.get("chief_complaint","").strip()
        history=request.form.get("history","").strip()
        examination=request.form.get("examination","").strip()
        diagnosis=request.form.get("diagnosis","").strip()
        treatment=request.form.get("treatment_plan","").strip()
        notes=request.form.get("notes","").strip()
        follow_up=request.form.get("follow_up_date","").strip()
        today_str=datetime.now().strftime("%Y-%m-%d")
        now_str=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("""INSERT INTO doctor_consultations
            (visit_id,patient_id,patient_name,opd_number,consult_date,doctor_name,
             chief_complaint,history,examination,diagnosis,treatment_plan,notes,status,follow_up_date)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,'Open',?)""",
            (visit_id,visit["patient_id"],patient["full_name"] if patient else "",
             visit["opd_number"],today_str,session.get("user"),
             complaint,history,examination,diagnosis,treatment,notes,follow_up or None))
        consult_id = c.lastrowid
        if action == "send_lab":
            tests_json=request.form.get("tests_json","[]")
            if tests_json and tests_json != "[]":
                c.execute("""INSERT INTO doctor_lab_requests
                    (consultation_id,patient_id,patient_name,tests_json,requested_by,requested_at,status)
                    VALUES(?,?,?,?,?,?,'Pending')""",
                    (consult_id,visit["patient_id"],patient["full_name"] if patient else "",
                     tests_json,session.get("user"),now_str))
                c.execute("UPDATE opd_visits SET lab_status='Pending', current_location='Lab' WHERE id=?", (visit_id,))
                flash("Lab request sent to Lab module.","info")
        if action == "prescribe":
            drugs_json=request.form.get("drugs_json","[]")
            if drugs_json and drugs_json != "[]":
                rx_no=generate_rx_no()
                try: drugs=json.loads(drugs_json); total=sum(float(d.get("total",0)) for d in drugs)
                except: total=0.0
                c.execute("""INSERT INTO prescriptions
                    (consultation_id,patient_id,patient_name,opd_number,prescribed_by,
                     prescribed_at,drugs_json,total_amount,invoice_no,status)
                    VALUES(?,?,?,?,?,?,?,?,?,'Pending')""",
                    (consult_id,visit["patient_id"],patient["full_name"] if patient else "",
                     visit["opd_number"],session.get("user"),now_str,drugs_json,total,rx_no))
                ph_inv=generate_invoice_no("PH")
                c.execute("""INSERT OR IGNORE INTO invoices
                    (invoice_no,invoice_type,patient_id,patient_name,opd_number,
                     items_json,total_amount,created_at,created_by)
                    VALUES(?,'Pharmacy',?,?,?,?,?,?,?)""",
                    (ph_inv,visit["patient_id"],patient["full_name"] if patient else "",
                     visit["opd_number"],drugs_json,total,now_str,session.get("user")))
                c.execute("UPDATE opd_visits SET pharmacy_status='Pending', current_location='Pharmacy' WHERE id=?", (visit_id,))
                flash(f"Prescription sent to Pharmacy. RX: {rx_no}","success")
        c.execute("UPDATE opd_visits SET doctor_status='Done' WHERE id=?", (visit_id,))
        conn.commit(); conn.close()
        log_activity(session.get("user"),"DOCTOR_CONSULT",f"Consulted {patient['full_name'] if patient else visit_id}")
        flash("Consultation saved.","success")
        return redirect(url_for("doctor"))
    conn.close()
    return render_template("doctor_consult.html", visit=visit, patient=patient, vitals=vitals,
                           prev_consults=prev_consults, lab_tests=lab_tests, drugs_list=drugs_list)


@app.route("/doctor/consultation/<int:cid>")
@login_required(role="Doctor")
def doctor_view_consult(cid):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM doctor_consultations WHERE id=?", (cid,))
    consult = c.fetchone()
    if not consult:
        conn.close(); flash("Not found.","danger"); return redirect(url_for("doctor"))
    c.execute("SELECT * FROM opd_patients WHERE patient_id=?", (consult["patient_id"],))
    patient = c.fetchone()
    c.execute("SELECT * FROM triage_records WHERE visit_id=? ORDER BY id DESC LIMIT 1", (consult["visit_id"],))
    vitals = c.fetchone()
    c.execute("SELECT * FROM doctor_lab_requests WHERE consultation_id=?", (cid,))
    lab_reqs = c.fetchall()
    c.execute("SELECT * FROM lab_requests WHERE patient_id=? ORDER BY id DESC LIMIT 5", (consult["patient_id"],))
    lab_results_list = c.fetchall()
    c.execute("SELECT * FROM prescriptions WHERE consultation_id=?", (cid,))
    prescriptions = c.fetchall()
    conn.close()
    return render_template("doctor_view_consult.html", consult=consult, patient=patient,
                           vitals=vitals, lab_reqs=lab_reqs, lab_results_list=lab_results_list,
                           prescriptions=prescriptions)


# ============================================================
#  UPGRADED CASHIER
# ============================================================

@app.route("/cashier/invoice/new", methods=["GET","POST"])
@login_required(role="Cashier")
def cashier_new_invoice():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM services ORDER BY name")
    services = c.fetchall()
    if request.method == "POST":
        patient_id=request.form.get("patient_id","").strip()
        patient_name=request.form.get("patient_name","").strip()
        opd_number=request.form.get("opd_number","").strip()
        items_json=request.form.get("items_json","[]")
        notes=request.form.get("notes","").strip()
        try: items=json.loads(items_json); total=sum(float(i.get("amount",0)) for i in items)
        except: items=[]; total=0.0
        inv_no=generate_invoice_no("MAN")
        c.execute("""INSERT INTO invoices
            (invoice_no,invoice_type,patient_id,patient_name,opd_number,
             items_json,total_amount,created_at,created_by,notes)
            VALUES(?,'Manual',?,?,?,?,?,?,?,?)""",
            (inv_no,patient_id,patient_name,opd_number,json.dumps(items),
             total,datetime.now().strftime("%Y-%m-%d %H:%M:%S"),session.get("user"),notes))
        conn.commit(); conn.close()
        flash(f"Invoice created: {inv_no}","success")
        return redirect(url_for("cashier_invoice_view", inv_no=inv_no))
    conn.close()
    return render_template("cashier_invoice_new.html", services=services)


@app.route("/cashier/invoice/<inv_no>")
@login_required(role="Cashier")
def cashier_invoice_view(inv_no):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM invoices WHERE invoice_no=?", (inv_no,))
    invoice = c.fetchone()
    conn.close()
    if not invoice:
        flash("Invoice not found.","danger"); return redirect(url_for("cashier"))
    try: items=json.loads(invoice["items_json"] or "[]")
    except: items=[]
    return render_template("cashier_invoice_view.html", invoice=invoice, items=items)


@app.route("/cashier/invoice/<inv_no>/pay", methods=["POST"])
@login_required(role="Cashier")
def cashier_pay_invoice(inv_no):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM invoices WHERE invoice_no=?", (inv_no,))
    invoice = c.fetchone()
    if not invoice:
        conn.close(); flash("Invoice not found.","danger"); return redirect(url_for("cashier"))
    payment_method=request.form.get("payment_method","Cash")
    receipt=generate_receipt_number()
    now_str=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today=datetime.now().strftime("%Y-%m-%d")
    c.execute("UPDATE invoices SET is_paid=1, paid_at=?, receipt_no=? WHERE invoice_no=?", (now_str,receipt,inv_no))
    amount_words=amount_to_words(invoice["total_amount"])
    c.execute("""INSERT INTO patient_records
        (date,receipt_number,patient_id,patient_name,service_received,
         amount_paid,amount_in_words,cashier_name,payment_method,details)
        VALUES(?,?,?,?,?,?,?,?,?,?)""",
        (today,receipt,invoice["patient_id"],invoice["patient_name"],
         invoice["invoice_type"],invoice["total_amount"],amount_words,
         session.get("user"),payment_method,f"Invoice {inv_no}"))
    rec_id=c.lastrowid
    itype=invoice["invoice_type"]
    if itype in ("Pharmacy","PH"):
        c.execute("UPDATE prescriptions SET is_paid=1 WHERE invoice_no=?", (inv_no,))
    if itype in ("Lab","INV"):
        c.execute("UPDATE lab_requests SET is_paid=1 WHERE invoice_no=?", (inv_no,))
    if itype == "Mortuary":
        c.execute("UPDATE mortuary_cases SET embalming_paid=1 WHERE corpse_id=?", (invoice["patient_id"],))
    conn.commit(); conn.close()
    log_activity(session.get("user"),"PAYMENT",f"Paid invoice {inv_no} receipt {receipt}")
    flash(f"Payment recorded. Receipt: {receipt}","success")
    return redirect(url_for("receipt_view", rec_id=rec_id))


@app.route("/cashier/pending-invoices")
@login_required(role="Cashier")
def cashier_pending_invoices():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM invoices WHERE is_paid=0 ORDER BY id DESC")
    invoices = c.fetchall()
    conn.close()
    return render_template("cashier_pending.html", invoices=invoices)


# ============================================================
#  UPGRADED LAB MODULE
# ============================================================

@app.route("/lab/requests")
@login_required(role="Lab")
def lab_requests_list():
    conn = get_db(); c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT * FROM doctor_lab_requests WHERE status='Pending' ORDER BY id DESC")
    doctor_requests = c.fetchall()
    c.execute("SELECT * FROM lab_requests WHERE substr(created_at,1,10)=? ORDER BY id DESC", (today,))
    today_labs = c.fetchall()
    c.execute("SELECT COUNT(*) FROM lab_requests WHERE substr(created_at,1,10)=?", (today,))
    today_count = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(total_amount),0) FROM lab_requests WHERE substr(created_at,1,10)=?", (today,))
    today_total = c.fetchone()[0]
    conn.close()
    return render_template("lab_requests.html", doctor_requests=doctor_requests,
        today_labs=today_labs, today_count=today_count, today_total=today_total, today=today)


@app.route("/lab/doctor-request/<int:req_id>", methods=["GET","POST"])
@login_required(role="Lab")
def lab_process_doctor_request(req_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM doctor_lab_requests WHERE id=?", (req_id,))
    dr_req = c.fetchone()
    if not dr_req:
        conn.close(); flash("Request not found.","danger"); return redirect(url_for("lab_requests_list"))
    c.execute("SELECT * FROM opd_patients WHERE patient_id=?", (dr_req["patient_id"],))
    patient = c.fetchone()
    c.execute("SELECT * FROM lab_test_prices ORDER BY name")
    all_tests = c.fetchall()
    if request.method == "POST":
        insured_status=request.form.get("insured_status","Non-insured")
        tests_json=request.form.get("tests_json","[]")
        try: total=float(request.form.get("total_amount","0"))
        except: total=0.0
        invoice=generate_lab_invoice()
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("""INSERT INTO lab_requests
            (created_at,patient_id,patient_name,age,insured_status,
             tests_json,total_amount,invoice_no,is_paid,results_json,results_at)
            VALUES(?,?,?,?,?,?,?,?,0,NULL,NULL)""",
            (created_at,dr_req["patient_id"],dr_req["patient_name"],
             patient["age"] if patient else "",insured_status,tests_json,total,invoice))
        lab_id=c.lastrowid
        inv_no=generate_invoice_no("LAB")
        c.execute("""INSERT OR IGNORE INTO invoices
            (invoice_no,invoice_type,patient_id,patient_name,opd_number,
             items_json,total_amount,created_at,created_by)
            VALUES(?,'Lab',?,?,?,?,?,?,?)""",
            (inv_no,dr_req["patient_id"],dr_req["patient_name"],
             patient["opd_number"] if patient else "",
             tests_json,total,created_at,session.get("user")))
        c.execute("UPDATE doctor_lab_requests SET status='Processing' WHERE id=?", (req_id,))
        conn.commit(); conn.close()
        log_activity(session.get("user"),"LAB_INVOICE",f"Invoice {invoice} for {dr_req['patient_name']}")
        flash(f"Lab invoice {invoice} created. Send patient to Cashier.","success")
        return redirect(url_for("lab_invoice", req_id=lab_id))
    conn.close()
    return render_template("lab_process_request.html", dr_req=dr_req, patient=patient, all_tests=all_tests)


@app.route("/lab/result/<int:req_id>/notify")
@login_required(role="Lab")
def lab_result_notify(req_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM lab_requests WHERE id=?", (req_id,))
    req = c.fetchone()
    if req:
        c.execute("UPDATE opd_visits SET lab_status='Done' WHERE patient_id=? AND lab_status='Pending'", (req["patient_id"],))
        conn.commit()
        log_activity(session.get("user"),"LAB_RESULT_READY",f"Results ready {req['patient_name']}")
        flash("Results marked ready. Doctor notified.","success")
    conn.close()
    return redirect(url_for("lab_results", req_id=req_id))


# ============================================================
#  PHARMACY MODULE
# ============================================================

@app.route("/pharmacy")
@login_required(role="Pharmacy")
def pharmacy():
    conn = get_db(); c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")
    c.execute("SELECT * FROM prescriptions WHERE status='Pending' ORDER BY id DESC")
    pending_rx = c.fetchall()
    c.execute("SELECT COUNT(*) FROM pharmacy_dispensing WHERE substr(dispensed_at,1,10)=?", (today,))
    today_dispensed = c.fetchone()[0]
    c.execute("SELECT COALESCE(SUM(total_amount),0) FROM pharmacy_dispensing WHERE substr(dispensed_at,1,10)=? AND is_paid=1", (today,))
    today_revenue = c.fetchone()[0]
    c.execute("SELECT * FROM pharmacy_drugs WHERE stock_qty <= reorder_level AND is_active=1 ORDER BY stock_qty ASC LIMIT 8")
    low_stock = c.fetchall()
    c.execute("SELECT COUNT(*) FROM pharmacy_drugs WHERE is_active=1")
    total_drugs = c.fetchone()[0]
    conn.close()
    return render_template("pharmacy.html", pending_rx=pending_rx,
        today_dispensed=today_dispensed, today_revenue=today_revenue,
        low_stock=low_stock, total_drugs=total_drugs, today=today)

@app.route('/pharmacy/drug/edit/<int:did>', methods=['GET', 'POST'])
def pharmacy_drug_edit(did):
    conn = sqlite3.connect('hospital.db')
    c = conn.cursor()

    if request.method == 'POST':
        drug_name = request.form['drug_name']
        generic_name = request.form['generic_name']
        category = request.form['category']
        unit = request.form['unit']
        unit_price = request.form['unit_price']
        stock_qty = request.form['stock_qty']
        reorder_level = request.form['reorder_level']

        c.execute("""
            UPDATE drugs SET
            drug_name=?, generic_name=?, category=?, unit=?,
            unit_price=?, stock_qty=?, reorder_level=?
            WHERE id=?
        """, (drug_name, generic_name, category, unit,
              unit_price, stock_qty, reorder_level, did))

        conn.commit()
        conn.close()

        flash("Drug updated successfully", "success")
        return redirect(url_for('pharmacy'))

    c.execute("SELECT * FROM drugs WHERE id=?", (did,))
    drug = c.fetchone()
    conn.close()

    return render_template("edit_drug.html", drug=drug)


@app.route("/pharmacy/dispense/<int:rx_id>", methods=["GET","POST"])
@login_required(role="Pharmacy")
def pharmacy_dispense(rx_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM prescriptions WHERE id=?", (rx_id,))
    rx = c.fetchone()
    if not rx:
        conn.close(); flash("Prescription not found.","danger"); return redirect(url_for("pharmacy"))
    if request.method == "POST":
        if not rx["is_paid"]:
            conn.close(); flash("Cannot dispense. Patient has not paid.","danger")
            return redirect(url_for("pharmacy_dispense", rx_id=rx_id))
        now_str=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        inv_no=generate_pharm_inv()
        c.execute("""INSERT INTO pharmacy_dispensing
            (prescription_id,patient_id,patient_name,dispensed_by,dispensed_at,
             drugs_json,total_amount,invoice_no,is_paid)
            VALUES(?,?,?,?,?,?,?,?,1)""",
            (rx_id,rx["patient_id"],rx["patient_name"],session.get("user"),now_str,
             rx["drugs_json"],rx["total_amount"],inv_no))
        try:
            drugs=json.loads(rx["drugs_json"] or "[]")
            for d in drugs:
                c.execute("UPDATE pharmacy_drugs SET stock_qty=stock_qty-? WHERE drug_name=? AND stock_qty>0",
                          (int(d.get("qty",1)),d.get("name","")))
        except: pass
        c.execute("UPDATE prescriptions SET status='Dispensed', is_dispensed=1 WHERE id=?", (rx_id,))
        conn.commit(); conn.close()
        log_activity(session.get("user"),"PHARMACY_DISPENSE",f"Dispensed to {rx['patient_name']}")
        flash("Drugs dispensed successfully!","success")
        return redirect(url_for("pharmacy"))
    try: drugs=json.loads(rx["drugs_json"] or "[]")
    except: drugs=[]
    conn.close()
    return render_template("pharmacy_dispense.html", rx=rx, drugs=drugs)


@app.route("/pharmacy/drugs")
@login_required(role="Pharmacy")
def pharmacy_drugs():
    conn = get_db(); c = conn.cursor()
    q=request.args.get("q","").strip(); cat=request.args.get("cat","").strip()
    if q:
        c.execute("SELECT * FROM pharmacy_drugs WHERE (drug_name LIKE ? OR generic_name LIKE ?) AND is_active=1 ORDER BY drug_name", (f"%{q}%",f"%{q}%"))
    elif cat:
        c.execute("SELECT * FROM pharmacy_drugs WHERE category=? AND is_active=1 ORDER BY drug_name", (cat,))
    else:
        c.execute("SELECT * FROM pharmacy_drugs WHERE is_active=1 ORDER BY drug_name")
    drugs = c.fetchall()
    c.execute("SELECT DISTINCT category FROM pharmacy_drugs WHERE is_active=1 AND category!='' ORDER BY category")
    categories = [r["category"] for r in c.fetchall()]
    conn.close()
    return render_template("pharmacy_drugs.html", drugs=drugs, q=q, cat=cat, categories=categories)


@app.route("/pharmacy/drug/add", methods=["GET","POST"])
@login_required(role="Pharmacy")
def pharmacy_drug_add():
    drug = None
    if request.method == "POST":
        conn = get_db(); c = conn.cursor()
        did = request.form.get("drug_id","")
        if did:
            c.execute("""UPDATE pharmacy_drugs SET drug_name=?,generic_name=?,category=?,unit=?,
                unit_price=?,stock_qty=?,reorder_level=?,supplier=?,expiry_date=? WHERE id=?""",
                (request.form.get("drug_name","").strip(), request.form.get("generic_name","").strip(),
                 request.form.get("category","").strip(), request.form.get("unit","Tablet"),
                 float(request.form.get("unit_price",0) or 0), int(request.form.get("stock_qty",0) or 0),
                 int(request.form.get("reorder_level",10) or 10), request.form.get("supplier","").strip(),
                 request.form.get("expiry_date","").strip(), did))
        else:
            c.execute("""INSERT OR IGNORE INTO pharmacy_drugs
                (drug_name,generic_name,category,unit,unit_price,stock_qty,reorder_level,supplier,expiry_date)
                VALUES(?,?,?,?,?,?,?,?,?)""",
                (request.form.get("drug_name","").strip(), request.form.get("generic_name","").strip(),
                 request.form.get("category","").strip(), request.form.get("unit","Tablet"),
                 float(request.form.get("unit_price",0) or 0), int(request.form.get("stock_qty",0) or 0),
                 int(request.form.get("reorder_level",10) or 10), request.form.get("supplier","").strip(),
                 request.form.get("expiry_date","").strip()))
        conn.commit(); conn.close()
        flash("Drug saved.","success")
        return redirect(url_for("pharmacy_drugs"))
    did = request.args.get("id","")
    if did:
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM pharmacy_drugs WHERE id=?", (did,))
        drug = c.fetchone(); conn.close()
    return render_template("pharmacy_drug_add.html", drug=drug)


@app.route("/pharmacy/upload-drugs", methods=["GET","POST"])
@login_required(role="Pharmacy")
def pharmacy_upload_drugs():
    if request.method == "POST":
        f = request.files.get("drug_file")
        if not f or not (f.filename.endswith(".xlsx") or f.filename.endswith(".xls") or f.filename.endswith(".csv")):
            flash("Upload an Excel (.xlsx/.xls) or CSV file.","danger")
            return redirect(url_for("pharmacy_upload_drugs"))
        import tempfile, os as _os
        suffix = _os.path.splitext(f.filename)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        f.save(tmp.name); tmp.close()
        try:
            if suffix == ".csv":
                import csv
                with open(tmp.name, newline='', encoding='utf-8-sig') as cf:
                    rows = list(csv.DictReader(cf))
            else:
                import openpyxl
                wb = openpyxl.load_workbook(tmp.name, data_only=True)
                ws = wb.active
                headers = [str(c2.value or "").strip().lower() for c2 in ws[1]]
                rows = [{headers[i]: (row[i] or "") for i in range(len(headers))} for row in ws.iter_rows(min_row=2, values_only=True)]
            conn = get_db(); c = conn.cursor()
            added = 0
            for row in rows:
                name=str(row.get("drug_name","") or row.get("name","") or row.get("drug","")).strip()
                if not name: continue
                try:
                    c.execute("""INSERT OR IGNORE INTO pharmacy_drugs
                        (drug_name,generic_name,category,unit,unit_price,stock_qty,reorder_level)
                        VALUES(?,?,?,?,?,?,?)""",
                        (name, str(row.get("generic_name","") or "").strip(),
                         str(row.get("category","") or "").strip(),
                         str(row.get("unit","") or "Tablet").strip(),
                         float(row.get("unit_price",0) or row.get("price",0) or 0),
                         int(float(row.get("stock_qty",0) or row.get("quantity",0) or row.get("qty",0) or 0)),
                         int(float(row.get("reorder_level",10) or 10))))
                    added += 1
                except: pass
            conn.commit(); conn.close(); _os.unlink(tmp.name)
            flash(f"✅ {added} drugs imported successfully!","success")
        except Exception as e:
            flash(f"Error: {e}","danger")
        return redirect(url_for("pharmacy_drugs"))
    return render_template("pharmacy_upload.html")


@app.route("/pharmacy/invoice/<int:rx_id>")
@login_required(role="Pharmacy")
def pharmacy_invoice(rx_id):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM prescriptions WHERE id=?", (rx_id,))
    rx = c.fetchone()
    if not rx:
        conn.close(); flash("Not found.","danger"); return redirect(url_for("pharmacy"))
    try: drugs=json.loads(rx["drugs_json"] or "[]")
    except: drugs=[]
    conn.close()
    return render_template("pharmacy_invoice.html", rx=rx, drugs=drugs)


if __name__ == "__main__":
    init_db()
    app.run(debug=True)
